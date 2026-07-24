#!/usr/bin/env python3
"""Descarga, normaliza y publica las series oficiales del observatorio.

Fuentes conectadas
------------------
* INE: población residente por comunidad autónoma.
* Seguridad Social: afiliación media por comunidad autónoma.
* SEPE: paro registrado nacional y por comunidad autónoma.
* Imserso: personas con derecho sin resolución de PIA y resoluciones de PIA.
* INE: PIB regional, PIB por habitante, EPA, renta, pobreza y vivienda turística.

El script conserva la última versión válida de una serie si una fuente falla. La
salida se escribe de forma atómica en ``data/dashboard.json``.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import io
import json
import os
import re
import tempfile
import unicodedata
import shutil
import subprocess
from dataclasses import asdict, dataclass
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Callable
from urllib.parse import urljoin
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from pypdf import PdfReader

from extended_collectors import (
    collect_dependency_extended,
    collect_health_waiting,
    collect_hospital_resources,
    collect_house_prices,
    collect_imv,
    collect_pension_reserve,
    collect_pensions,
    collect_primary_care_centers,
    collect_residences,
    collect_university_system,
    derive_education_series,
    infer_category,
)

ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / "data" / "dashboard.json"
MADRID = ZoneInfo("Europe/Madrid")
USER_AGENT = "ObservatorioServiciosPublicos/1.0 (reutilizacion estadistica; contacto academico)"

INE_POPULATION_CSV = "https://www.ine.es/jaxiT3/files/t/csv_bdsc/67217.csv"
INE_POPULATION_PAGE = "https://www.ine.es/jaxiT3/Tabla.htm?t=67217"
INE_EPA_UNEMPLOYED_CSV = "https://www.ine.es/jaxiT3/files/t/es/csv_bdsc/65332.csv"
INE_EPA_UNEMPLOYED_PAGE = "https://www.ine.es/jaxiT3/Tabla.htm?t=65332"
INE_EPA_RATE_CSV = "https://www.ine.es/jaxiT3/files/t/es/csv_bdsc/65334.csv"
INE_EPA_RATE_PAGE = "https://www.ine.es/jaxiT3/Tabla.htm?t=65334"
INE_INCOME_CSV = "https://www.ine.es/jaxiT3/files/t/es/csv_bdsc/9947.csv"
INE_INCOME_PAGE = "https://www.ine.es/jaxiT3/Tabla.htm?t=9947"
INE_POVERTY_CSV = "https://www.ine.es/jaxiT3/files/t/es/csv_bdsc/9963.csv"
INE_POVERTY_PAGE = "https://www.ine.es/jaxiT3/Tabla.htm?t=9963"
INE_TOURIST_DWELLINGS_CSV = "https://www.ine.es/jaxiT3/files/t/es/csv_bd/46141.csv"
INE_TOURIST_DWELLINGS_PAGE = "https://www.ine.es/jaxiT3/Tabla.htm?t=46141"
INE_REGIONAL_ACCOUNTS_PAGE = "https://www.ine.es/dyngs/INEbase/es/operacion.htm?c=Estadistica_C&cid=1254736167628&menu=resultados&idp=1254735576581"
INE_REGIONAL_ACCOUNTS_FALLBACK = "https://www.ine.es/daco/daco42/cre00/b2024/pr_cre.xlsx"
SEPE_NATIONAL_XLS = "https://www.sepe.es/SiteSepe/contenidos/que_es_el_sepe/estadisticas/datos_avance/xls/empleo/evolparo.xls"
SEPE_REGIONAL_XLS = "https://www.sepe.es/SiteSepe/contenidos/que_es_el_sepe/estadisticas/datos_avance/xls/empleo/evolprovsectores.xls"
SEPE_PAGE = "https://www.sepe.es/HomeSepe/que-es-el-sepe/estadisticas/datos-avance/paro.html"
IMSERSO_MONTHLY_PAGE = "https://imserso.es/el-imserso/documentacion/estadisticas/sistema-autonomia-atencion-dependencia-saad/estadisticas-mensual"
SOCIAL_SECURITY_REPORTS_PAGE = "https://www.seg-social.es/wps/portal/wss/internet/EstadisticasPresupuestosEstudios/Estadisticas/EST8/f32c1896-f56d-4728-a4ac-fa0b410ea0b2/42105a1e-060c-47f4-9b3c-33c530a606ca"

EDUCATION_BASE = "https://estadisticas.educacion.gob.es/EducaJaxiPx"
EDUCATION_TABLES = [
    ("general", "Enseñanzas de régimen general", "alumnado_1_01", "Alumnado total de enseñanzas de régimen general"),
    ("infant", "Educación Infantil", "alumnado_2_01", "Alumnado de Educación Infantil"),
    ("early_childhood", "Primer ciclo de Educación Infantil (0-3)", "alumnado_2_03", "Alumnado de primer ciclo de Educación Infantil (0-3 años)"),
    ("primary", "Educación Primaria", "alumnado_3_01", "Alumnado de Educación Primaria"),
    ("eso", "Educación Secundaria Obligatoria", "alumnado_4_01", "Alumnado de Educación Secundaria Obligatoria"),
    ("bachiller", "Bachillerato", "alumnado_5_01", "Alumnado de Bachillerato"),
    ("fp_basic", "FP de Grado Básico", "alumnado_6_01", "Alumnado de ciclos formativos de FP de Grado Básico"),
    ("fp_middle", "FP de Grado Medio", "alumnado_6_03", "Alumnado de ciclos formativos de FP de Grado Medio"),
    ("fp_higher", "FP de Grado Superior", "alumnado_6_11", "Alumnado de ciclos formativos de FP de Grado Superior"),
]
EDUCATION_CSV_TEMPLATE = EDUCATION_BASE + "/files/_px/es/csv_bdsc/no-universitaria/alumnado/matriculado/series-new/gen-alumnado/l0/{table}.csv_bdsc"
EDUCATION_PAGE_TEMPLATE = EDUCATION_BASE + "/Tabla.htm?L=0&file={table}.px&path=%2Fno-universitaria%2Falumnado%2Fmatriculado%2Fseries-new%2Fgen-alumnado%2Fl0%2F"


@dataclass
class SourceCheck:
    source_id: str
    ok: bool
    checked_at: str
    message: str
    checksum: str | None = None
    source_url: str | None = None


class LinkCollector(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self._href: str | None = None
        self._text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() == "a":
            self._href = dict(attrs).get("href")
            self._text = []

    def handle_data(self, data: str) -> None:
        if self._href is not None:
            self._text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "a" and self._href is not None:
            self.links.append((self._href, " ".join(self._text).strip()))
            self._href = None
            self._text = []


def fetch_bytes(url: str, timeout: int = 120) -> bytes:
    request = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "*/*"})
    with urlopen(request, timeout=timeout) as response:
        return response.read()


def sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", delete=False, dir=path.parent) as tmp:
        json.dump(payload, tmp, ensure_ascii=False, indent=2)
        tmp.write("\n")
        temp_name = tmp.name
    os.replace(temp_name, path)


def normalize_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(c for c in text if not unicodedata.combining(c)).upper().strip()


def parse_spanish_number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if text in {"", ".", "..", "...", "-", "nan", "None"}:
        raise ValueError(f"Valor numérico ausente: {value!r}")
    text = text.replace(".", "").replace(",", ".")
    return float(text)


def try_spanish_number(value: Any) -> float | None:
    try:
        return parse_spanish_number(value)
    except (TypeError, ValueError):
        return None


def make_points(pairs: list[tuple[str, float | int]]) -> list[dict[str, Any]]:
    seen: dict[str, float | int] = {}
    for period, value in pairs:
        seen[str(period)] = value
    return [{"period": period, "value": seen[period]} for period in sorted(seen)]


TERRITORIES: dict[str, tuple[str, str]] = {
    "ESPANA": ("ES", "España"),
    "TOTAL NACIONAL": ("ES", "España"),
    "TOTAL": ("ES", "España"),
    "ANDALUCIA": ("ES-AN", "Andalucía"),
    "ARAGON": ("ES-AR", "Aragón"),
    "ASTURIAS, PRINCIPADO DE": ("ES-AS", "Asturias"),
    "PRINCIPADO DE ASTURIAS": ("ES-AS", "Asturias"),
    "ASTURIAS": ("ES-AS", "Asturias"),
    "BALEARS, ILLES": ("ES-IB", "Illes Balears"),
    "ILLES BALEARS": ("ES-IB", "Illes Balears"),
    "BALEARES": ("ES-IB", "Illes Balears"),
    "CANARIAS": ("ES-CN", "Canarias"),
    "CANTABRIA": ("ES-CB", "Cantabria"),
    "CASTILLA Y LEON": ("ES-CL", "Castilla y León"),
    "CASTILLA - LA MANCHA": ("ES-CM", "Castilla-La Mancha"),
    "CASTILLA-LA MANCHA": ("ES-CM", "Castilla-La Mancha"),
    "CASTILLA LA MANCHA": ("ES-CM", "Castilla-La Mancha"),
    "CATALUNA": ("ES-CT", "Cataluña"),
    "COMUNITAT VALENCIANA": ("ES-VC", "Comunitat Valenciana"),
    "COM. VALENCIANA": ("ES-VC", "Comunitat Valenciana"),
    "CMDAD. VALENCIANA": ("ES-VC", "Comunitat Valenciana"),
    "COMUNIDAD VALENCIANA": ("ES-VC", "Comunitat Valenciana"),
    "EXTREMADURA": ("ES-EX", "Extremadura"),
    "GALICIA": ("ES-GA", "Galicia"),
    "MADRID, COMUNIDAD DE": ("ES-MD", "Comunidad de Madrid"),
    "COM. DE MADRID": ("ES-MD", "Comunidad de Madrid"),
    "CMDAD. DE MADRID": ("ES-MD", "Comunidad de Madrid"),
    "COMUNIDAD DE MADRID": ("ES-MD", "Comunidad de Madrid"),
    "MURCIA, REGION DE": ("ES-MC", "Región de Murcia"),
    "REGION DE MURCIA": ("ES-MC", "Región de Murcia"),
    "NAVARRA, COMUNIDAD FORAL DE": ("ES-NC", "Comunidad Foral de Navarra"),
    "COM. FORAL DE NAVARRA": ("ES-NC", "Comunidad Foral de Navarra"),
    "NAVARRA": ("ES-NC", "Comunidad Foral de Navarra"),
    "PAIS VASCO": ("ES-PV", "País Vasco"),
    "RIOJA, LA": ("ES-RI", "La Rioja"),
    "LA RIOJA": ("ES-RI", "La Rioja"),
    "CEUTA": ("ES-CE", "Ceuta"),
    "MELILLA": ("ES-ML", "Melilla"),
}


def territory(value: Any) -> tuple[str, str] | None:
    key = normalize_key(value)
    key = re.sub(r"^\d+\s+", "", key)
    key = re.sub(r"\s+\(\d+\)$", "", key)
    key = key.replace("NAVARRA (COMUNIDAD FORAL DE)", "NAVARRA, COMUNIDAD FORAL DE")
    return TERRITORIES.get(key)


def order_territories(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(items, key=lambda item: (item["code"] != "ES", item["name"]))


def local_or_fetch(local_dir: Path | None, filename: str, url: str) -> tuple[bytes, str]:
    if local_dir:
        candidate = local_dir / filename
        if candidate.exists():
            return candidate.read_bytes(), url
    return fetch_bytes(url), url


def series_base(
    *,
    id_: str,
    name: str,
    description: str,
    unit: str,
    frequency: str,
    source: str,
    source_url: str,
    status: str = "Oficial",
    decimals: int = 0,
) -> dict[str, Any]:
    return {
        "id": id_,
        "name": name,
        "description": description,
        "unit": unit,
        "frequency": frequency,
        "source": source,
        "source_url": source_url,
        "status": status,
        "decimals": decimals,
        "territories": [],
    }


def collect_population(local_dir: Path | None = None) -> tuple[dict[str, Any], SourceCheck]:
    content, fetched_from = local_or_fetch(local_dir, "ine67217.csv", INE_POPULATION_CSV)
    text = content.decode("utf-8-sig")
    rows = csv.DictReader(io.StringIO(text), delimiter=";")
    values: dict[tuple[str, str], list[tuple[str, int]]] = {}
    for row in rows:
        if row.get("Sexo") != "Total" or row.get("Edad") != "Todas las edades" or row.get("Nacionalidad") != "Total":
            continue
        if (row.get("Provincias") or "").strip():
            continue
        raw_name = (row.get("Comunidades y Ciudades Autónomas") or "").strip() or "Total Nacional"
        raw_name = re.sub(r"^\d+\s+", "", raw_name)
        terr = territory(raw_name)
        if not terr:
            continue
        code, name = terr
        period = str(row["Periodo"]).strip()
        value = int(parse_spanish_number(row["Total"]))
        values.setdefault((code, name), []).append((period, value))

    series = series_base(
        id_="population",
        name="Población residente",
        description="Población residente a 1 de enero, según la Estadística Continua de Población.",
        unit="personas",
        frequency="Anual",
        source="INE · Estadística Continua de Población",
        source_url=INE_POPULATION_PAGE,
    )
    series["territories"] = order_territories([
        {"code": code, "name": name, "points": make_points(points)}
        for (code, name), points in values.items()
    ])
    check = SourceCheck("ine_population", True, datetime.now(MADRID).isoformat(timespec="seconds"), f"{len(series['territories'])} territorios; {sum(len(t['points']) for t in series['territories'])} observaciones", sha256(content), fetched_from)
    return series, check


def discover_social_security_pdf() -> str:
    page = fetch_bytes(SOCIAL_SECURITY_REPORTS_PAGE).decode("utf-8", errors="replace")
    parser = LinkCollector()
    parser.feed(page)
    candidates: list[str] = []
    for href, text in parser.links:
        joined = f"{href} {text}".lower()
        if "afiliados" in joined and "ccaa" in joined and ".pdf" in joined:
            candidates.append(urljoin(SOCIAL_SECURITY_REPORTS_PAGE, html.unescape(href)))
    if not candidates:
        # Algunas páginas insertan el enlace como texto escapado fuera de un ancla estándar.
        regex = re.findall(r'href=["\']([^"\']*Afiliados[^"\']*CCAA[^"\']*\.pdf[^"\']*)', page, flags=re.I)
        candidates.extend(urljoin(SOCIAL_SECURITY_REPORTS_PAGE, html.unescape(x)) for x in regex)
    if not candidates:
        raise RuntimeError("No se encontró el informe mensual de afiliados por CCAA")
    # Los nombres terminan habitualmente en -AAAAMM.pdf. Escogemos el periodo mayor.
    return max(candidates, key=lambda url: re.findall(r"(20\d{4})", url)[-1] if re.findall(r"(20\d{4})", url) else "000000")


def collect_affiliation(local_dir: Path | None = None) -> tuple[dict[str, Any], SourceCheck]:
    if local_dir and (local_dir / "afiliados_ccaa_202606.pdf").exists():
        pdf_url = SOCIAL_SECURITY_REPORTS_PAGE
        content = (local_dir / "afiliados_ccaa_202606.pdf").read_bytes()
        fetched_from = SOCIAL_SECURITY_REPORTS_PAGE
        year, month = 2026, 6
    else:
        pdf_url = discover_social_security_pdf()
        content = fetch_bytes(pdf_url)
        fetched_from = pdf_url
        matches = re.findall(r"(20\d{2})(0[1-9]|1[0-2])", pdf_url)
        if not matches:
            raise RuntimeError("No se pudo identificar el periodo del informe de afiliación")
        year, month = map(int, matches[-1])

    reader = PdfReader(io.BytesIO(content))
    table_text = ""
    for page in reader.pages[:12]:
        text = page.extract_text() or ""
        normalized = normalize_key(text)
        if (
            "AFILIADOS MEDIOS POR COMUNIDADES AUTONOMAS" in normalized
            and "TOTAL SISTEMA" in normalized
            and re.search(r"ANDALUCIA\s+\d[\d.]*,\d{2}", normalized)
        ):
            table_text = text
            break
    if not table_text:
        raise RuntimeError("No se localizó la tabla de afiliación por CCAA en el PDF")

    current_period = f"{year:04d}-{month:02d}"
    previous_month_date = date(year if month > 1 else year - 1, month - 1 if month > 1 else 12, 1)
    previous_period = f"{previous_month_date.year:04d}-{previous_month_date.month:02d}"
    previous_year_period = f"{year - 1:04d}-{month:02d}"
    points_by_territory: dict[tuple[str, str], list[tuple[str, float]]] = {}

    known_names = sorted({key for key in TERRITORIES if key not in {"ESPANA", "TOTAL NACIONAL"}}, key=len, reverse=True)
    for raw_line in table_text.splitlines():
        line_key = normalize_key(raw_line)
        matched_name = next((name for name in known_names if line_key == name or line_key.startswith(name + " ")), None)
        if not matched_name:
            continue
        terr = TERRITORIES[matched_name]
        # El orden es: valor actual, diferencia mensual, % mensual,
        # diferencia interanual, % interanual, valor del mismo mes del año anterior…
        numbers = re.findall(r"-?\d[\d.]*,\d{2}", raw_line)
        if len(numbers) < 6:
            continue
        current = parse_spanish_number(numbers[0])
        monthly_difference = parse_spanish_number(numbers[1])
        previous_year = parse_spanish_number(numbers[5])
        previous_month = current - monthly_difference
        points_by_territory[terr] = [
            (previous_year_period, round(previous_year, 2)),
            (previous_period, round(previous_month, 2)),
            (current_period, round(current, 2)),
        ]

    if ("ES", "España") not in points_by_territory:
        # La fila TOTAL no siempre se reconoce mediante el mapa de nombres.
        total_line = next((line for line in table_text.splitlines() if normalize_key(line).startswith("TOTAL ")), "")
        numbers = re.findall(r"-?\d[\d.]*,\d{2}", total_line)
        if len(numbers) >= 6:
            current = parse_spanish_number(numbers[0])
            monthly_difference = parse_spanish_number(numbers[1])
            previous_year = parse_spanish_number(numbers[5])
            points_by_territory[("ES", "España")] = [
                (previous_year_period, round(previous_year, 2)),
                (previous_period, round(current - monthly_difference, 2)),
                (current_period, round(current, 2)),
            ]

    series = series_base(
        id_="social_security_affiliates",
        name="Afiliación media a la Seguridad Social",
        description="Promedio mensual de personas afiliadas en alta laboral en el total del sistema.",
        unit="afiliaciones medias",
        frequency="Mensual",
        source="Tesorería General de la Seguridad Social",
        source_url=pdf_url,
        decimals=2,
    )
    series["territories"] = order_territories([
        {"code": code, "name": name, "points": make_points(points)}
        for (code, name), points in points_by_territory.items()
    ])
    if len(series["territories"]) < 18:
        raise RuntimeError(f"Solo se han reconocido {len(series['territories'])} territorios en afiliación")
    check = SourceCheck("social_security_affiliation", True, datetime.now(MADRID).isoformat(timespec="seconds"), f"Informe {current_period}; {len(series['territories'])} territorios", sha256(content), fetched_from)
    return series, check


def xls_bytes_to_workbook(content: bytes, stem: str) -> Any:
    """Convierte el XLS oficial a XLSX con LibreOffice y lo abre con openpyxl."""
    executable = shutil.which("libreoffice") or shutil.which("soffice")
    if not executable:
        mac_soffice = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
        executable = str(mac_soffice) if mac_soffice.exists() else None
    if not executable:
        raise RuntimeError(
            "Para leer los XLS del SEPE se necesita LibreOffice. "
            "En GitHub Actions se instala automáticamente."
        )
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp = Path(tmp_dir)
        source = tmp / f"{stem}.xls"
        source.write_bytes(content)
        process = subprocess.run(
            [executable, "--headless", "--convert-to", "xlsx", "--outdir", str(tmp), str(source)],
            capture_output=True,
            text=True,
            timeout=120,
            check=False,
        )
        output = tmp / f"{stem}.xlsx"
        if process.returncode != 0 or not output.exists():
            raise RuntimeError(f"LibreOffice no pudo convertir {stem}.xls: {process.stderr.strip()}")
        # Cargar en memoria antes de borrar el directorio temporal.
        converted = output.read_bytes()
    return load_workbook(io.BytesIO(converted), data_only=True, read_only=True)


def parse_sepe_national(content: bytes) -> list[tuple[str, int]]:
    workbook = xls_bytes_to_workbook(content, "evolparo")
    sheet = workbook.active
    year_columns: list[tuple[int, int]] = []
    for row in range(1, min(9, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            value = normalize_key(sheet.cell(row, col).value)
            match = re.search(r"ANO\s+(20\d{2})", value)
            if match:
                year_columns.append((col, int(match.group(1))))
    if not year_columns:
        raise RuntimeError("No se identificaron las columnas anuales del fichero nacional de paro")

    result: list[tuple[str, int]] = []
    for row in range(1, sheet.max_row + 1):
        month_number = sheet.cell(row, 1).value
        if not isinstance(month_number, (int, float)) or int(month_number) not in range(1, 13):
            continue
        for col, year in year_columns:
            value = sheet.cell(row, col).value
            if isinstance(value, (int, float)) and value > 0:
                result.append((f"{year}-{int(month_number):02d}", int(value)))
    return sorted(result)


def parse_sepe_regional(content: bytes) -> tuple[str, dict[tuple[str, str], list[tuple[str, int]]]]:
    workbook = xls_bytes_to_workbook(content, "evolprovsectores")
    sheet = workbook["Total sectores"] if "Total sectores" in workbook.sheetnames else workbook.active
    heading = " ".join(str(sheet.cell(r, 1).value or "") for r in range(1, min(7, sheet.max_row + 1)))
    match = re.search(r"(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)\s+DE\s+(20\d{2})", normalize_key(heading))
    if not match:
        raise RuntimeError("No se identificó el periodo del fichero regional de paro")
    months = {name: i for i, name in enumerate(["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"], 1)}
    month = months[match.group(1)]
    year = int(match.group(2))
    current_period = f"{year}-{month:02d}"
    prev = date(year if month > 1 else year - 1, month - 1 if month > 1 else 12, 1)
    previous_period = f"{prev.year}-{prev.month:02d}"

    data: dict[tuple[str, str], list[tuple[str, int]]] = {}
    for row in range(1, sheet.max_row + 1):
        raw_name = sheet.cell(row, 2).value
        terr = territory(raw_name)
        if not terr:
            continue
        previous = sheet.cell(row, 3).value
        current = sheet.cell(row, 4).value
        if isinstance(previous, (int, float)) and isinstance(current, (int, float)):
            data[terr] = [(previous_period, int(previous)), (current_period, int(current))]
    return current_period, data

def collect_unemployment(local_dir: Path | None = None) -> tuple[dict[str, Any], SourceCheck]:
    national_content, national_from = local_or_fetch(local_dir, "evolparo.xls", SEPE_NATIONAL_XLS)
    regional_content, regional_from = local_or_fetch(local_dir, "evolprovsectores.xls", SEPE_REGIONAL_XLS)
    national_points = parse_sepe_national(national_content)
    current_period, regional = parse_sepe_regional(regional_content)
    regional[("ES", "España")] = national_points

    series = series_base(
        id_="registered_unemployment",
        name="Paro registrado",
        description="Personas inscritas como demandantes de empleo clasificadas como paradas en los servicios públicos de empleo.",
        unit="personas",
        frequency="Mensual",
        source="Servicio Público de Empleo Estatal (SEPE)",
        source_url=SEPE_PAGE,
    )
    series["territories"] = order_territories([
        {"code": code, "name": name, "points": make_points(points)}
        for (code, name), points in regional.items()
    ])
    check = SourceCheck("sepe_registered_unemployment", True, datetime.now(MADRID).isoformat(timespec="seconds"), f"Periodo {current_period}; {len(series['territories'])} territorios", sha256(national_content + regional_content), f"{national_from}; {regional_from}")
    return series, check


def discover_imserso_xlsx() -> str:
    page = fetch_bytes(IMSERSO_MONTHLY_PAGE).decode("utf-8", errors="replace")
    parser = LinkCollector()
    parser.feed(page)
    candidates: list[str] = []
    for href, text in parser.links:
        joined = f"{href} {text}".lower()
        if "estsisaad_" in joined and ".xlsx" in joined:
            candidates.append(urljoin(IMSERSO_MONTHLY_PAGE, html.unescape(href)))
    if not candidates:
        regex = re.findall(r'["\']([^"\']*estsisaad_\d{8}\.xlsx[^"\']*)["\']', page, flags=re.I)
        candidates.extend(urljoin(IMSERSO_MONTHLY_PAGE, html.unescape(x)) for x in regex)
    if not candidates:
        raise RuntimeError("No se encontró el último Excel mensual del SAAD")
    return max(candidates, key=lambda url: re.findall(r"estsisaad_(\d{8})", url)[-1] if re.findall(r"estsisaad_(\d{8})", url) else "00000000")


def normalize_excel_period(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None

def parse_dependency_sheet(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    ws = workbook[sheet_name]
    period_columns: list[tuple[int, str]] = []
    for col in range(4, min(ws.max_column, 20) + 1):
        value = ws.cell(7, col).value
        if value:
            period = normalize_excel_period(value)
            if period:
                period_columns.append((col, period))
    territories: list[dict[str, Any]] = []
    for row in range(9, ws.max_row + 1):
        raw_name = ws.cell(row, 2).value
        terr = territory(raw_name)
        if not terr:
            continue
        code, name = terr
        points: list[tuple[str, int]] = []
        for col, period in period_columns:
            value = ws.cell(row, col).value
            if isinstance(value, (int, float)):
                points.append((period, int(value)))
        if points:
            territories.append({"code": code, "name": name, "points": make_points(points)})
    return order_territories(territories)


def collect_dependency(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    if local_dir and (local_dir / "estsisaad_20260630.xlsx").exists():
        source_url = IMSERSO_MONTHLY_PAGE
        content = (local_dir / "estsisaad_20260630.xlsx").read_bytes()
        fetched_from = IMSERSO_MONTHLY_PAGE
    else:
        source_url = discover_imserso_xlsx()
        content = fetch_bytes(source_url)
        fetched_from = source_url
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    waiting = series_base(
        id_="dependency_waiting_pia",
        name="Personas con derecho a prestación sin resolución de PIA",
        description="Personas con grado de dependencia reconocido que todavía no disponen de una resolución de su Programa Individual de Atención.",
        unit="personas",
        frequency="Mensual (serie con cortes seleccionados)",
        source="Imserso · Sistema para la Autonomía y Atención a la Dependencia",
        source_url=source_url,
    )
    waiting["territories"] = parse_dependency_sheet(workbook, "EVO_sinPIA")

    pia = series_base(
        id_="dependency_pia_resolutions",
        name="Resoluciones de Programa Individual de Atención (PIA)",
        description="Evolución del número de resoluciones de PIA por comunidad autónoma.",
        unit="resoluciones",
        frequency="Mensual (serie con cortes seleccionados)",
        source="Imserso · Sistema para la Autonomía y Atención a la Dependencia",
        source_url=source_url,
    )
    pia["territories"] = parse_dependency_sheet(workbook, "EVO_resolPIA")
    if len(waiting["territories"]) < 18 or len(pia["territories"]) < 18:
        raise RuntimeError("El Excel SAAD no contiene todos los territorios esperados")
    latest = waiting["territories"][0]["points"][-1]["period"]
    check = SourceCheck("imserso_saad", True, datetime.now(MADRID).isoformat(timespec="seconds"), f"Corte {latest}; dos series y {len(waiting['territories'])} territorios", sha256(content), fetched_from)
    return [waiting, pia], check



def discover_regional_accounts_xlsx() -> str:
    """Localiza el Excel de principales resultados de la revisión más reciente."""
    try:
        page = fetch_bytes(INE_REGIONAL_ACCOUNTS_PAGE).decode("utf-8", errors="replace")
        parser = LinkCollector()
        parser.feed(page)
        candidates = [
            urljoin(INE_REGIONAL_ACCOUNTS_PAGE, html.unescape(href))
            for href, text in parser.links
            if "pr_cre.xlsx" in f"{href} {text}".lower()
        ]
        if candidates:
            return candidates[0]
        regex = re.findall(r'["\']([^"\']*pr_cre\.xlsx[^"\']*)["\']', page, flags=re.I)
        if regex:
            return urljoin(INE_REGIONAL_ACCOUNTS_PAGE, html.unescape(regex[0]))
    except Exception:
        pass
    return INE_REGIONAL_ACCOUNTS_FALLBACK


def parse_regional_accounts_sheet(
    workbook: Any,
    sheet_name: str,
    *,
    transform: Callable[[float], float | int],
) -> list[dict[str, Any]]:
    ws = workbook[sheet_name]
    year_columns: list[tuple[int, str]] = []
    for col in range(3, ws.max_column + 1):
        value = ws.cell(5, col).value
        match = re.search(r"(20\d{2})", str(value or ""))
        if match:
            year_columns.append((col, match.group(1)))
    if not year_columns:
        raise RuntimeError(f"No se encontraron años en {sheet_name}")

    values: dict[tuple[str, str], list[tuple[str, float | int]]] = {}
    for row in range(7, ws.max_row + 1):
        terr = territory(ws.cell(row, 2).value)
        if not terr:
            continue
        points: list[tuple[str, float | int]] = []
        for col, period in year_columns:
            raw = ws.cell(row, col).value
            if isinstance(raw, (int, float)):
                points.append((period, transform(float(raw))))
        if points:
            values[terr] = points
    return order_territories([
        {"code": code, "name": name, "points": make_points(points)}
        for (code, name), points in values.items()
    ])


def collect_regional_accounts(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    if local_dir and (local_dir / "pr_cre.xlsx").exists():
        content = (local_dir / "pr_cre.xlsx").read_bytes()
        fetched_from = INE_REGIONAL_ACCOUNTS_FALLBACK
    else:
        source_url = discover_regional_accounts_xlsx()
        content = fetch_bytes(source_url)
        fetched_from = source_url
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    gdp = series_base(
        id_="regional_gdp",
        name="Producto interior bruto regional",
        description="PIB a precios de mercado y precios corrientes. Los valores de 2023 son provisionales y los de 2024 son avance en la edición conectada.",
        unit="millones de euros",
        frequency="Anual",
        source="INE · Contabilidad Regional de España",
        source_url=INE_REGIONAL_ACCOUNTS_PAGE,
        status="Oficial · últimos años provisionales/avance",
        decimals=1,
    )
    gdp["territories"] = parse_regional_accounts_sheet(
        workbook, "Tabla_1", transform=lambda value: round(value / 1000, 1)
    )

    gdp_pc = series_base(
        id_="regional_gdp_per_capita",
        name="PIB por habitante",
        description="Producto interior bruto regional por habitante a precios corrientes.",
        unit="euros por habitante",
        frequency="Anual",
        source="INE · Contabilidad Regional de España",
        source_url=INE_REGIONAL_ACCOUNTS_PAGE,
        status="Oficial · últimos años provisionales/avance",
    )
    gdp_pc["territories"] = parse_regional_accounts_sheet(
        workbook, "Tabla_2", transform=lambda value: int(round(value))
    )
    if len(gdp["territories"]) < 18 or len(gdp_pc["territories"]) < 18:
        raise RuntimeError("La Contabilidad Regional no contiene todos los territorios esperados")
    latest = gdp["territories"][0]["points"][-1]["period"]
    check = SourceCheck(
        "ine_regional_accounts", True, datetime.now(MADRID).isoformat(timespec="seconds"),
        f"Serie hasta {latest}; PIB y PIB por habitante para {len(gdp['territories'])} territorios",
        sha256(content), fetched_from,
    )
    return [gdp, gdp_pc], check


def parse_ine_csv(content: bytes, *, delimiter: str = ";") -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text), delimiter=delimiter))


def collect_epa(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    unemployed_content, unemployed_from = local_or_fetch(local_dir, "ine65332.csv", INE_EPA_UNEMPLOYED_CSV)
    rate_content, rate_from = local_or_fetch(local_dir, "ine65334.csv", INE_EPA_RATE_CSV)

    def extract(content: bytes, *, multiply: float = 1.0) -> dict[tuple[str, str], list[tuple[str, float | int]]]:
        values: dict[tuple[str, str], list[tuple[str, float | int]]] = {}
        for row in parse_ine_csv(content):
            if row.get("Sexo") != "Ambos sexos" or row.get("Edad") != "Total":
                continue
            terr = territory(row.get("Comunidades y Ciudades Autónomas"))
            raw = try_spanish_number(row.get("Total"))
            if not terr or raw is None:
                continue
            value: float | int = raw * multiply
            if multiply == 1000:
                value = int(round(value))
            else:
                value = round(value, 2)
            values.setdefault(terr, []).append((str(row.get("Periodo")), value))
        return values

    unemployed_values = extract(unemployed_content, multiply=1000)
    rate_values = extract(rate_content)

    unemployed = series_base(
        id_="epa_unemployed",
        name="Personas paradas según la EPA",
        description="Estimación trimestral de personas paradas de ambos sexos y todas las edades. Los valores proceden de una encuesta muestral.",
        unit="personas",
        frequency="Trimestral",
        source="INE · Encuesta de Población Activa",
        source_url=INE_EPA_UNEMPLOYED_PAGE,
        status="Oficial · estimación muestral",
    )
    unemployed["territories"] = order_territories([
        {"code": code, "name": name, "points": make_points(points)}
        for (code, name), points in unemployed_values.items()
    ])

    rate = series_base(
        id_="epa_unemployment_rate",
        name="Tasa de paro según la EPA",
        description="Porcentaje de población activa que se encuentra parada, para ambos sexos y todas las edades.",
        unit="porcentaje",
        frequency="Trimestral",
        source="INE · Encuesta de Población Activa",
        source_url=INE_EPA_RATE_PAGE,
        status="Oficial · estimación muestral",
        decimals=2,
    )
    rate["symbol"] = " %"
    rate["territories"] = order_territories([
        {"code": code, "name": name, "points": make_points(points)}
        for (code, name), points in rate_values.items()
    ])
    if len(unemployed["territories"]) < 18 or len(rate["territories"]) < 18:
        raise RuntimeError("Las tablas EPA no contienen todos los territorios esperados")
    latest = unemployed["territories"][0]["points"][-1]["period"]
    check = SourceCheck(
        "ine_epa", True, datetime.now(MADRID).isoformat(timespec="seconds"),
        f"Serie hasta {latest}; parados y tasa para {len(unemployed['territories'])} territorios",
        sha256(unemployed_content + rate_content), f"{unemployed_from}; {rate_from}",
    )
    return [unemployed, rate], check


def collect_income_and_poverty(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    income_content, income_from = local_or_fetch(local_dir, "ine9947.csv", INE_INCOME_CSV)
    poverty_content, poverty_from = local_or_fetch(local_dir, "ine9963.csv", INE_POVERTY_CSV)

    income_values: dict[tuple[str, str], list[tuple[str, int]]] = {}
    for row in parse_ine_csv(income_content):
        if row.get("Renta anual neta media por persona y por unidad de consumo") != "Renta neta media por persona":
            continue
        terr = territory(row.get("Comunidades y Ciudades Autónomas"))
        raw = try_spanish_number(row.get("Total"))
        if terr and raw is not None:
            income_values.setdefault(terr, []).append((str(row.get("Periodo")), int(round(raw))))

    poverty_label = "Tasa de riesgo de pobreza (renta del año anterior a la entrevista)"
    poverty_values: dict[tuple[str, str], list[tuple[str, float]]] = {}
    for row in parse_ine_csv(poverty_content):
        if row.get("Tasa de riesgo de pobreza") != poverty_label:
            continue
        terr = territory(row.get("Comunidades y Ciudades Autónomas"))
        raw = try_spanish_number(row.get("Total"))
        if terr and raw is not None:
            poverty_values.setdefault(terr, []).append((str(row.get("Periodo")), round(raw, 1)))

    income = series_base(
        id_="mean_net_income_per_person",
        name="Renta neta media por persona",
        description="Ingreso neto anual medio por persona. La renta se refiere al año anterior al de la entrevista de la ECV.",
        unit="euros por persona",
        frequency="Anual",
        source="INE · Encuesta de Condiciones de Vida",
        source_url=INE_INCOME_PAGE,
        status="Oficial · estimación muestral",
    )
    income["territories"] = order_territories([
        {"code": code, "name": name, "points": make_points(points)}
        for (code, name), points in income_values.items()
    ])

    poverty = series_base(
        id_="poverty_risk_rate",
        name="Tasa de riesgo de pobreza",
        description="Porcentaje de personas con ingresos por unidad de consumo inferiores al 60 % de la mediana.",
        unit="porcentaje",
        frequency="Anual",
        source="INE · Encuesta de Condiciones de Vida",
        source_url=INE_POVERTY_PAGE,
        status="Oficial · estimación muestral",
        decimals=1,
    )
    poverty["symbol"] = " %"
    poverty["territories"] = order_territories([
        {"code": code, "name": name, "points": make_points(points)}
        for (code, name), points in poverty_values.items()
    ])
    if len(income["territories"]) < 18 or len(poverty["territories"]) < 18:
        raise RuntimeError("Las tablas de renta y pobreza no contienen todos los territorios esperados")
    latest = income["territories"][0]["points"][-1]["period"]
    check = SourceCheck(
        "ine_income_poverty", True, datetime.now(MADRID).isoformat(timespec="seconds"),
        f"ECV hasta {latest}; renta y pobreza para {len(income['territories'])} territorios",
        sha256(income_content + poverty_content), f"{income_from}; {poverty_from}",
    )
    return [income, poverty], check


def collect_tourist_dwellings(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    content, fetched_from = local_or_fetch(local_dir, "ine46141.csv", INE_TOURIST_DWELLINGS_CSV)
    counts: dict[tuple[str, str], list[tuple[str, int]]] = {}
    shares: dict[tuple[str, str], list[tuple[str, float]]] = {}
    for row in parse_ine_csv(content, delimiter="\t"):
        if (row.get("Provincias") or "").strip():
            continue
        raw_geo = (row.get("Comunidades y Ciudades Autónomas") or "").strip()
        if not raw_geo and (row.get("Total Nacional") or "").strip():
            raw_geo = "Total Nacional"
        terr = territory(raw_geo)
        if not terr:
            continue
        raw = try_spanish_number(row.get("Total"))
        if raw is None:
            continue
        period = str(row.get("Periodo"))
        measure = row.get("Viviendas y plazas")
        if measure == "Viviendas turísticas":
            counts.setdefault(terr, []).append((period, int(round(raw))))
        elif measure == "Porcentaje de viviendas turísticas sobre el total de viviendas censadas":
            shares.setdefault(terr, []).append((period, round(raw, 2)))

    count_series = series_base(
        id_="tourist_dwellings",
        name="Viviendas destinadas a uso turístico",
        description="Estimación del número de viviendas turísticas ofrecidas en las principales plataformas digitales.",
        unit="viviendas",
        frequency="Semestral / cortes extraordinarios",
        source="INE · Estadística experimental de viviendas turísticas",
        source_url=INE_TOURIST_DWELLINGS_PAGE,
        status="Estadística experimental",
    )
    count_series["territories"] = order_territories([
        {"code": code, "name": name, "points": make_points(points)}
        for (code, name), points in counts.items()
    ])

    share_series = series_base(
        id_="tourist_dwellings_share",
        name="Peso de las viviendas turísticas en el parque residencial",
        description="Porcentaje estimado de viviendas turísticas sobre el total de viviendas censadas.",
        unit="porcentaje",
        frequency="Semestral / cortes extraordinarios",
        source="INE · Estadística experimental de viviendas turísticas",
        source_url=INE_TOURIST_DWELLINGS_PAGE,
        status="Estadística experimental",
        decimals=2,
    )
    share_series["symbol"] = " %"
    share_series["territories"] = order_territories([
        {"code": code, "name": name, "points": make_points(points)}
        for (code, name), points in shares.items()
    ])
    if len(count_series["territories"]) < 18 or len(share_series["territories"]) < 18:
        raise RuntimeError("La tabla de vivienda turística no contiene todos los territorios esperados")
    latest = count_series["territories"][0]["points"][-1]["period"]
    check = SourceCheck(
        "ine_tourist_dwellings", True, datetime.now(MADRID).isoformat(timespec="seconds"),
        f"Serie hasta {latest}; recuento y porcentaje para {len(count_series['territories'])} territorios",
        sha256(content), fetched_from,
    )
    return [count_series, share_series], check


def collect_education_enrolment(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    """Recoge alumnado público y privado por etapa y comunidad autónoma."""
    output: list[dict[str, Any]] = []
    checksums: list[str] = []
    source_urls: list[str] = []
    total_observations = 0

    for key, label, table, description in EDUCATION_TABLES:
        csv_url = EDUCATION_CSV_TEMPLATE.format(table=table)
        page_url = EDUCATION_PAGE_TEMPLATE.format(table=table)
        local_name = f"educa_{table}.csv"
        content, fetched_from = local_or_fetch(local_dir, local_name, csv_url)
        checksums.append(sha256(content))
        source_urls.append(fetched_from)
        rows = csv.DictReader(io.StringIO(content.decode("utf-8-sig")), delimiter=";")
        values: dict[str, dict[tuple[str, str], list[tuple[str, int]]]] = {
            "public": {},
            "private": {},
        }
        for row in rows:
            ownership = normalize_key(row.get("Titularidad del centro"))
            ownership_key = {"CENTROS PUBLICOS": "public", "CENTROS PRIVADOS": "private"}.get(ownership)
            if not ownership_key:
                continue
            terr = territory(row.get("Comunidad autónoma"))
            if not terr:
                continue
            raw_value = try_spanish_number(row.get("Total"))
            if raw_value is None:
                continue
            period = str(row.get("periodo") or "").strip()
            if not period:
                continue
            values[ownership_key].setdefault(terr, []).append((period, int(round(raw_value))))
            total_observations += 1

        for ownership_key, ownership_label in (("public", "centros públicos"), ("private", "centros privados")):
            series = series_base(
                id_=f"education_{key}_{ownership_key}",
                name=f"{label}: alumnado en {ownership_label}",
                description=(
                    f"{description}, desglosado por titularidad del centro. "
                    + ("La categoría privada incluye centros concertados y no concertados; esta serie no los separa." if ownership_key == "private" else "La categoría pública se refiere a centros de titularidad pública.")
                ),
                unit="alumnos",
                frequency="Curso académico",
                source="EDUCAbase · Ministerio de Educación, Formación Profesional y Deportes",
                source_url=page_url,
                status="Oficial",
            )
            series["territories"] = order_territories([
                {"code": code, "name": name, "points": make_points(points)}
                for (code, name), points in values[ownership_key].items()
            ])
            if len(series["territories"]) < 18:
                raise RuntimeError(f"{table}: solo se reconocieron {len(series['territories'])} territorios para {ownership_label}")
            output.append(series)

    combined_checksum = sha256("".join(checksums).encode("ascii"))
    latest = max(
        (territory_item["points"][-1]["period"] for series in output for territory_item in series["territories"] if territory_item["points"]),
        default="—",
    )
    check = SourceCheck(
        "education_enrolment",
        True,
        datetime.now(MADRID).isoformat(timespec="seconds"),
        f"{len(output)} series; {total_observations} observaciones; último curso {latest}",
        combined_checksum,
        "https://www.educacionfpydeportes.gob.es/servicios-al-ciudadano/estadisticas/no-universitaria/alumnado/matriculado/series.html",
    )
    return output, check


def period_sort_key(period: str) -> tuple[int, int, int]:
    text = str(period)
    if match := re.fullmatch(r"(20\d{2})-(\d{2})-(\d{2})", text):
        return tuple(map(int, match.groups()))
    if match := re.fullmatch(r"(20\d{2})-(\d{2})", text):
        year, month = map(int, match.groups())
        return year, month, 31
    if match := re.fullmatch(r"(20\d{2})M(\d{2})", text):
        year, month = map(int, match.groups())
        return year, month, 31
    if match := re.fullmatch(r"(20\d{2})[-]?Q([1-4])", text):
        year, quarter = map(int, match.groups())
        return year, quarter * 3, 31
    if match := re.fullmatch(r"(20\d{2})T([1-4])", text):
        year, quarter = map(int, match.groups())
        return year, quarter * 3, 31
    if match := re.fullmatch(r"(20\d{2})", text):
        return int(match.group(1)), 12, 31
    return 0, 0, 0


Collector = Callable[[Path | None], tuple[Any, SourceCheck]]


def update_catalog_connections(catalog: list[dict[str, Any]], series_list: list[dict[str, Any]]) -> None:
    """Vincula el catálogo con las series disponibles y añade las nuevas.

    Se conservan las fichas conceptuales originales, aunque algunas aún no tengan
    una fuente homogénea. Las series detalladas obtenidas de las fuentes se añaden
    automáticamente al catálogo para que sean accesibles con un clic.
    """
    available = {series["id"]: series for series in series_list}
    by_name = {series["name"]: series["id"] for series in series_list}
    aliases = {
        "Población residente": "population",
        "Afiliación a la Seguridad Social": "social_security_affiliates",
        "PIB nacional y regional": "regional_gdp",
        "PIB por habitante": "regional_gdp_per_capita",
        "Renta neta media por persona": "mean_net_income_per_person",
        "Parados según EPA": "epa_unemployed",
        "Tasa de paro según EPA": "epa_unemployment_rate",
        "Paro registrado": "registered_unemployment",
        "Tasa de riesgo de pobreza": "poverty_risk_rate",
        "Viviendas destinadas a uso turístico": "tourist_dwellings",
        "Peso de las viviendas turísticas en el parque residencial": "tourist_dwellings_share",
        "Personas pendientes de resolución o prestación": "dependency_waiting_pia",
        "Personas con resolución de PIA": "dependency_pia_resolutions",
        "Tiempo de tramitación de la dependencia": "dependency_wait_request_to_benefit",
        "Número de pensionistas": "pensioners_total",
        "Pensiones de jubilación": "retirement_pensions",
        "Evolución de la nómina de pensiones": "pensions_monthly_payroll",
        "Evolución de la Caja de las pensiones": "pension_reserve_fund",
        "Estado de la hucha de las pensiones": "pension_reserve_fund",
        "Fondo de Reserva de la Seguridad Social": "pension_reserve_fund",
        "Universidades públicas y privadas": "universities_total",
        "Precio medio del metro cuadrado construido": "house_appraised_value_m2",
        "Beneficiarios del Ingreso Mínimo Vital": "imv_beneficiaries",
        "Listas de espera quirúrgicas por especialidad": "surgical_waiting_patients",
        "Centros de salud y consultorios del SNS": "primary_care_centers_total",
        "Primeras consultas por especialidad": "consultation_waiting_patients",
        "Número de residencias públicas y privadas": "residential_centers_total",
        "Plazas públicas, concertadas y privadas": "residential_places_total",
        "Coste o precio por plaza": "residential_public_price",
        "Número de residentes": "residential_users",
    }
    seen_ids: set[str] = set()
    for item in catalog:
        item.pop("series_id", None)
        series_id = aliases.get(item.get("name")) or by_name.get(item.get("name"))
        if series_id in available:
            item["series_id"] = series_id
            item["connected"] = True
            seen_ids.add(series_id)
        else:
            item["connected"] = False

    existing_names = {item.get("name") for item in catalog}
    for series in series_list:
        if series["id"] in seen_ids or series["name"] in existing_names:
            continue
        catalog.append({
            "category": infer_category(series["id"]),
            "name": series["name"],
            "availability": "high",
            "automation": "automatic",
            "frequency": series.get("frequency", "—"),
            "source": series.get("source", "Fuente oficial"),
            "summary": series.get("description", "Serie oficial conectada al observatorio."),
            "series_id": series["id"],
            "connected": True,
        })
        existing_names.add(series["name"])
        seen_ids.add(series["id"])


def update_dashboard(local_dir: Path | None = None) -> dict[str, Any]:
    with DATA_FILE.open(encoding="utf-8") as file:
        dashboard = json.load(file)
    old_series = {series["id"]: series for series in dashboard.get("series", [])}
    current_series: dict[str, dict[str, Any]] = {}
    checks: list[SourceCheck] = []

    def run_one(source_id: str, function: Collector) -> None:
        try:
            result, check = function(local_dir)
            if isinstance(result, list):
                for series in result:
                    current_series[series["id"]] = series
            else:
                current_series[result["id"]] = result
            checks.append(check)
        except Exception as exc:  # Cada fuente falla de forma independiente.
            now = datetime.now(MADRID).isoformat(timespec="seconds")
            checks.append(SourceCheck(source_id, False, now, f"{type(exc).__name__}: {exc}"))

    run_one("ine_population", collect_population)
    run_one("ine_regional_accounts", collect_regional_accounts)
    run_one("ine_epa", collect_epa)
    run_one("ine_income_poverty", collect_income_and_poverty)
    run_one("social_security_affiliation", collect_affiliation)
    run_one("sepe_registered_unemployment", collect_unemployment)
    run_one("imserso_saad", collect_dependency_extended)
    run_one("ine_tourist_dwellings", collect_tourist_dwellings)
    run_one("education_enrolment", collect_education_enrolment)
    run_one("pensions", collect_pensions)
    run_one("imv", collect_imv)
    run_one("health_waiting", collect_health_waiting)
    run_one("residences", collect_residences)
    run_one("mivau_house_price", collect_house_prices)
    run_one("pension_reserve", collect_pension_reserve)
    run_one("university_system", collect_university_system)
    run_one("hospital_catalog", collect_hospital_resources)
    run_one("primary_care_catalog", collect_primary_care_centers)

    # Las cuotas y totales educativos se recalculan sobre las series nuevas o,
    # si EDUCAbase falla, sobre la última copia válida conservada.
    education_inputs = dict(old_series)
    education_inputs.update(current_series)
    for series in derive_education_series(education_inputs):
        current_series[series["id"]] = series

    desired_order = [
        "population",
        "regional_gdp",
        "regional_gdp_per_capita",
        "mean_net_income_per_person",
        "epa_unemployed",
        "epa_unemployment_rate",
        "social_security_affiliates",
        "registered_unemployment",
        "poverty_risk_rate",
        "dependency_waiting_pia",
        "dependency_pia_resolutions",
        "tourist_dwellings",
        "tourist_dwellings_share",
        "education_general_public",
        "education_general_private",
        "education_infant_public",
        "education_infant_private",
        "education_early_childhood_public",
        "education_early_childhood_private",
        "education_primary_public",
        "education_primary_private",
        "education_eso_public",
        "education_eso_private",
        "education_bachiller_public",
        "education_bachiller_private",
        "education_fp_basic_public",
        "education_fp_basic_private",
        "education_fp_middle_public",
        "education_fp_middle_private",
        "education_fp_higher_public",
        "education_fp_higher_private",
    ]
    final_series: list[dict[str, Any]] = []
    used: set[str] = set()
    combined = dict(old_series)
    combined.update(current_series)
    for series_id in desired_order:
        series = combined.get(series_id)
        if series:
            final_series.append(series); used.add(series_id)
    # Publica también todas las series adicionales sin exigir que cada nueva
    # incorporación modifique manualmente la lista de orden.
    for series_id, series in sorted(combined.items(), key=lambda item: (infer_category(item[0]), item[1].get("name", ""))):
        if series_id not in used:
            final_series.append(series); used.add(series_id)
    dashboard["series"] = final_series
    update_catalog_connections(dashboard.setdefault("catalog", []), final_series)

    now = datetime.now(MADRID)
    latest_periods = [territory_item["points"][-1]["period"] for series in final_series for territory_item in series.get("territories", []) if territory_item.get("points")]
    latest_period = max(latest_periods, key=lambda p: (period_sort_key(p), len(str(p)))) if latest_periods else "—"
    signature = hashlib.sha256(json.dumps(final_series, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()[:12]
    dashboard["meta"] = {
        "title": "Observatorio de servicios públicos",
        "dataset_version": f"{now:%Y%m%d-%H%M}-{signature}",
        "last_checked": now.isoformat(timespec="seconds"),
        "latest_period": latest_period,
        "demo": False,
        "connected_series": len(final_series),
        "source_checks": [asdict(check) for check in checks],
        "notice": f"Versión funcional ampliada: {len(final_series)} series oficiales o calculadas a partir de datos oficiales. Las fuentes se actualizan de forma independiente.",
    }
    atomic_write_json(DATA_FILE, dashboard)
    return dashboard


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--local-source-dir",
        type=Path,
        help="Usa copias locales de las fuentes cuando existan; útil para pruebas reproducibles.",
    )
    # Compatibilidad con la versión anterior y el workflow ya publicado.
    parser.add_argument("--live-checks", action="store_true", help=argparse.SUPPRESS)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    dashboard = update_dashboard(args.local_source_dir)
    checks = dashboard["meta"].get("source_checks", [])
    failures = [check for check in checks if not check["ok"]]
    print(f"Actualización terminada: {dashboard['meta']['last_checked']}")
    print(f"Series publicadas: {len(dashboard.get('series', []))}")
    print(f"Fuentes comprobadas: {len(checks)}; fallos: {len(failures)}")
    for check in checks:
        print(f"- {'OK' if check['ok'] else 'ERROR'} {check['source_id']}: {check['message']}")
    # Solo se devuelve error si no queda ninguna serie publicable.
    if not dashboard.get("series"):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
