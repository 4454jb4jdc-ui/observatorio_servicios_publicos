"""Recolectores adicionales del Observatorio de Servicios Públicos.

El módulo mantiene independientes las fuentes más complejas (PDF y libros Excel
con estructuras propias) para que un fallo en una de ellas no afecte al resto.
"""
from __future__ import annotations

import csv
import hashlib
import html
import io
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urljoin
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from pypdf import PdfReader

MADRID = ZoneInfo("Europe/Madrid")
USER_AGENT = "ObservatorioServiciosPublicos/1.1 (reutilizacion estadistica)"

PENSIONS_PAGE = "https://www.seg-social.es/wps/portal/wss/internet/EstadisticasPresupuestosEstudios/Estadisticas/est23/est24"
IMV_STATS_PAGE = "https://www.inclusion.gob.es/web/inclusion/estadisticas-informes"
IMV_LATEST_PAGE = "https://www.inclusion.gob.es/w/el-ingreso-minimo-vital-llega-en-junio-a-mas-de-2.6-millones-de-personas-que-residen-en-860.458-hogares"
HEALTH_WAITING_PAGE = "https://www.sanidad.gob.es/estadEstudios/estadisticas/inforRecopilaciones/listaEspera.htm"
RESIDENCES_PAGE = "https://imserso.es/el-imserso/documentacion/estadisticas/servicios-sociales-dirigidos-a-personas-mayores-en-espana-diciembre-2024"
IMSERSO_MONTHLY_PAGE = "https://imserso.es/el-imserso/documentacion/estadisticas/sistema-autonomia-atencion-dependencia-saad/estadisticas-mensual"
HOUSE_PRICE_CSV = "https://datos.canarias.es/api/estadisticas/statistical-resources/v1.0/datasets/ISTAC/E20007A_000002/1.26.csv"
HOUSE_PRICE_PAGE = "https://datos.gob.es/es/catalogo/e05233601-valor-tasado-de-la-vivienda"
UNIVERSITY_REPORT_URL = "https://www.ciencia.gob.es/dam/jcr%3A747444d8-7089-4b15-97e1-0fc27874e6a7/DatosCifras_SUE2025_2026.pdf"
UNIVERSITY_PAGE = "https://www.ciencia.gob.es/Ministerio/Estadisticas/SIIU.html"
PENSION_RESERVE_URL = "https://revista.seg-social.es/-/el-fondo-de-reserva-de-la-seguridad-social-supera-los-15.200-millones-de-euros-en-abril"
HOSPITAL_CATALOG_XLSX = "https://www.sanidad.gob.es/estadEstudios/estadisticas/sisInfSanSNS/ofertaRecursos/hospitales/docs/CNH_2025.xlsx"
HOSPITAL_CATALOG_PAGE = "https://www.sanidad.gob.es/estadEstudios/estadisticas/sisInfSanSNS/ofertaRecursos/hospitales/home.htm"
PRIMARY_CARE_CATALOG_XLSX = "https://www.sanidad.gob.es/estadEstudios/estadisticas/docs/siap/2026_C_Catal_Centros_AP.xlsx"
PRIMARY_CARE_CATALOG_PAGE = "https://www.sanidad.gob.es/ciudadanos/prestaciones/centrosServiciosSNS/centrosSalud/home.htm"


@dataclass
class SourceCheck:
    source_id: str
    ok: bool
    checked_at: str
    message: str
    checksum: str | None = None
    source_url: str | None = None


class LinkCollector(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self._href: str | None = None
        self._text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() == "a":
            self._href = dict(attrs).get("href")
            self._text = []

    def handle_data(self, data: str) -> None:
        if self._href is not None:
            self._text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "a" and self._href is not None:
            self.links.append((self._href, " ".join(self._text).strip()))
            self._href = None
            self._text = []


def fetch_bytes(url: str, timeout: int = 120) -> bytes:
    request = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "*/*"})
    with urlopen(request, timeout=timeout) as response:
        return response.read()


def fetch_text(url: str) -> str:
    return fetch_bytes(url).decode("utf-8", errors="replace")


def sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def normalize_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", text.upper()).strip()


def parse_spanish_number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip().replace("\u00a0", "").replace("€", "").replace("%", "").replace(" ", "")
    if text in {"", "-", ".", "..", "...", "NONE", "NAN"}:
        raise ValueError("valor ausente")
    text = text.replace(".", "").replace(",", ".")
    return float(text)


def try_number(value: Any) -> float | None:
    try:
        return parse_spanish_number(value)
    except (TypeError, ValueError):
        return None


def make_points(pairs: Iterable[tuple[str, float | int]]) -> list[dict[str, Any]]:
    values: dict[str, float | int] = {}
    for period, value in pairs:
        values[str(period)] = value
    return [{"period": period, "value": values[period]} for period in sorted(values)]


TERRITORIES: dict[str, tuple[str, str]] = {
    "ESPANA": ("ES", "España"), "TOTAL": ("ES", "España"), "TOTAL NACIONAL": ("ES", "España"),
    "TOTAL SISTEMA": ("ES", "España"),
    "ANDALUCIA": ("ES-AN", "Andalucía"), "ARAGON": ("ES-AR", "Aragón"),
    "ASTURIAS": ("ES-AS", "Asturias"), "ASTURIAS (PRINCIPADO DE)": ("ES-AS", "Asturias"),
    "ASTURIAS, PRINCIPADO DE": ("ES-AS", "Asturias"), "PRINCIPADO DE ASTURIAS": ("ES-AS", "Asturias"),
    "PDO DE ASTURIAS": ("ES-AS", "Asturias"),
    "BALEARES": ("ES-IB", "Illes Balears"), "BALEARS (ILLES)": ("ES-IB", "Illes Balears"),
    "BALEARS, ILLES": ("ES-IB", "Illes Balears"), "ILLES BALEARS": ("ES-IB", "Illes Balears"),
    "CANARIAS": ("ES-CN", "Canarias"), "CANTABRIA": ("ES-CB", "Cantabria"),
    "CASTILLA Y LEON": ("ES-CL", "Castilla y León"),
    "CASTILLA - LA MANCHA": ("ES-CM", "Castilla-La Mancha"), "CASTILLA-LA MANCHA": ("ES-CM", "Castilla-La Mancha"),
    "CATALUNA": ("ES-CT", "Cataluña"), "CATALUNYA": ("ES-CT", "Cataluña"),
    "COMUNITAT VALENCIANA": ("ES-VC", "Comunitat Valenciana"), "COMUNIDAD VALENCIANA": ("ES-VC", "Comunitat Valenciana"),
    "C. VALENCIANA": ("ES-VC", "Comunitat Valenciana"), "COM. VALENCIANA": ("ES-VC", "Comunitat Valenciana"),
    "EXTREMADURA": ("ES-EX", "Extremadura"), "GALICIA": ("ES-GA", "Galicia"),
    "MADRID": ("ES-MD", "Comunidad de Madrid"), "MADRID (COM. DE)": ("ES-MD", "Comunidad de Madrid"),
    "MADRID (COMUNIDAD DE)": ("ES-MD", "Comunidad de Madrid"), "MADRID, COMUNIDAD DE": ("ES-MD", "Comunidad de Madrid"),
    "COMUNIDAD DE MADRID": ("ES-MD", "Comunidad de Madrid"),
    "MURCIA": ("ES-MC", "Región de Murcia"), "MURCIA (REGION DE)": ("ES-MC", "Región de Murcia"),
    "MURCIA, REGION DE": ("ES-MC", "Región de Murcia"), "REGION DE MURCIA": ("ES-MC", "Región de Murcia"),
    "NAVARRA": ("ES-NC", "Comunidad Foral de Navarra"), "NAVARRA (C. FORAL DE)": ("ES-NC", "Comunidad Foral de Navarra"),
    "NAVARRA (COMUNIDAD FORAL DE)": ("ES-NC", "Comunidad Foral de Navarra"),
    "NAVARRA, C. FORAL DE": ("ES-NC", "Comunidad Foral de Navarra"),
    "NAVARRA, COMUNIDAD FORAL DE": ("ES-NC", "Comunidad Foral de Navarra"),
    "C. FORAL DE NAVARRA": ("ES-NC", "Comunidad Foral de Navarra"),
    "PAIS VASCO": ("ES-PV", "País Vasco"), "RIOJA": ("ES-RI", "La Rioja"),
    "RIOJA (LA)": ("ES-RI", "La Rioja"), "RIOJA, LA": ("ES-RI", "La Rioja"), "LA RIOJA": ("ES-RI", "La Rioja"),
    "CEUTA": ("ES-CE", "Ceuta"), "MELILLA": ("ES-ML", "Melilla"),
}

# De mayor a menor longitud para no confundir, por ejemplo, Madrid con Comunidad de Madrid.
ALIASES = sorted(TERRITORIES, key=len, reverse=True)


def territory(value: Any) -> tuple[str, str] | None:
    key = normalize_key(value)
    key = re.sub(r"^\d+\s+", "", key)
    return TERRITORIES.get(key)


def match_line_territory(line: str) -> tuple[tuple[str, str], str] | None:
    key = normalize_key(line)
    for alias in ALIASES:
        if key == alias or key.startswith(alias + " "):
            # La longitud con acentos puede diferir. Se corta usando palabras del alias.
            n_words = len(alias.split())
            words = line.strip().split()
            remainder = " ".join(words[n_words:])
            # Los paréntesis y abreviaturas alteran el nº de palabras: localizamos el primer número.
            number_match = re.search(r"(?<!\w)-?\d", line)
            if number_match:
                remainder = line[number_match.start():]
            return TERRITORIES[alias], remainder
    return None


def order_territories(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(items, key=lambda item: (item["code"] != "ES", item["name"]))


def series_base(*, id_: str, name: str, description: str, unit: str, frequency: str,
                source: str, source_url: str, status: str = "Oficial", decimals: int = 0) -> dict[str, Any]:
    return {
        "id": id_, "name": name, "description": description, "unit": unit,
        "frequency": frequency, "source": source, "source_url": source_url,
        "status": status, "decimals": decimals, "territories": [],
    }


def single_period_series(*, id_: str, name: str, description: str, unit: str, frequency: str,
                         source: str, source_url: str, period: str,
                         values: dict[tuple[str, str], float | int | None], decimals: int = 0,
                         status: str = "Oficial") -> dict[str, Any]:
    series = series_base(id_=id_, name=name, description=description, unit=unit,
                         frequency=frequency, source=source, source_url=source_url,
                         status=status, decimals=decimals)
    series["territories"] = order_territories([
        {"code": code, "name": label, "points": [{"period": period, "value": value}]}
        for (code, label), value in values.items() if value is not None
    ])
    return series


def local_or_discover(local_dir: Path | None, filename: str, discover: callable) -> tuple[bytes, str]:
    if local_dir and (local_dir / filename).exists():
        return (local_dir / filename).read_bytes(), discover.__dict__.get("page", "") or ""
    url = discover()
    return fetch_bytes(url), url


def discover_link(page_url: str, predicate: callable, *, rank_pattern: str = r"(20\d{4,6})") -> str:
    page = fetch_text(page_url)
    parser = LinkCollector(); parser.feed(page)
    candidates: list[str] = []
    for href, text in parser.links:
        joined = f"{href} {text}"
        if predicate(joined):
            candidates.append(urljoin(page_url, html.unescape(href)))
    if not candidates:
        raise RuntimeError(f"No se localizó el fichero esperado en {page_url}")
    def rank(url: str) -> str:
        matches = re.findall(rank_pattern, url)
        return matches[-1] if matches else "00000000"
    return max(candidates, key=rank)


# ----------------------------- Pensiones ---------------------------------

def discover_pensions_pdf() -> str:
    return discover_link(
        PENSIONS_PAGE,
        lambda joined: ".pdf" in joined.lower() and "comunidades" in normalize_key(joined) and "PROVINCIAS" in normalize_key(joined),
    )

discover_pensions_pdf.page = PENSIONS_PAGE


def extract_numbers(text: str, *, preserve_missing: bool = False) -> list[float | None]:
    pattern = r"(?<!\w)-|[-+]?\d[\d.]*,\d+|[-+]?\d[\d.]*"
    result: list[float | None] = []
    for token in re.findall(pattern, text):
        if token == "-":
            if preserve_missing:
                result.append(None)
            continue
        try:
            result.append(parse_spanish_number(token))
        except ValueError:
            if preserve_missing:
                result.append(None)
    return result


def collect_pensions(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    content, fetched_from = local_or_discover(local_dir, "pensions_latest.pdf", discover_pensions_pdf)
    reader = PdfReader(io.BytesIO(content))
    title = "\n".join((reader.pages[i].extract_text() or "") for i in range(min(2, len(reader.pages))))
    month_names = {name: i for i, name in enumerate(
        ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"], 1)}
    match = re.search(r"(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)\s+(20\d{2})", normalize_key(title))
    if not match:
        raise RuntimeError("No se pudo identificar el periodo del boletín de pensiones")
    month, year = month_names[match.group(1)], int(match.group(2))
    period = f"{year:04d}-{month:02d}"

    rows: dict[tuple[str, str], list[float | None]] = {}
    for page in reader.pages[:25]:
        for line in (page.extract_text() or "").splitlines():
            matched = match_line_territory(line)
            if not matched:
                continue
            terr, remainder = matched
            numbers = extract_numbers(remainder)
            if len(numbers) >= 18 and terr not in rows:
                rows[terr] = numbers
    if len(rows) < 18:
        raise RuntimeError(f"Solo se reconocieron {len(rows)} territorios en el cuadro de pensiones")

    metric_specs = [
        ("pensions_total", "Pensiones contributivas en vigor", "Número total de pensiones contributivas en vigor del sistema de la Seguridad Social.", "pensiones", 0, 0),
        ("pensions_monthly_payroll", "Nómina mensual de pensiones contributivas", "Importe mensual agregado de las pensiones contributivas en vigor.", "euros al mes", 1, 2),
        ("pension_average", "Pensión media del sistema", "Importe medio mensual de las pensiones contributivas en vigor.", "euros al mes", 2, 2),
        ("retirement_pensions", "Pensiones de jubilación en vigor", "Número de pensiones contributivas de jubilación en vigor.", "pensiones", 6, 0),
        ("retirement_monthly_payroll", "Nómina mensual de pensiones de jubilación", "Importe mensual agregado de las pensiones de jubilación.", "euros al mes", 7, 2),
        ("retirement_pension_average", "Pensión media de jubilación", "Importe medio mensual de las pensiones contributivas de jubilación.", "euros al mes", 8, 2),
    ]
    output: list[dict[str, Any]] = []
    for id_, name, description, unit, index, decimals in metric_specs:
        values = {terr: (round(nums[index], decimals) if nums[index] is not None else None) for terr, nums in rows.items()}
        output.append(single_period_series(
            id_=id_, name=name, description=description, unit=unit, frequency="Mensual",
            source="INSS · Estadística de pensiones contributivas", source_url=PENSIONS_PAGE,
            period=period, values=values, decimals=decimals,
        ))

    # Pensionistas: incluye cuatro cortes regionales y una serie nacional mensual/anual.
    pensioner_points: dict[tuple[str, str], list[tuple[str, int]]] = {}
    for page in reader.pages[max(0, len(reader.pages) - 8):]:
        for line in (page.extract_text() or "").splitlines():
            matched = match_line_territory(line)
            if not matched:
                continue
            terr, remainder = matched
            nums = extract_numbers(remainder)
            if len(nums) >= 11:
                current, previous_month, previous_year, year10, year15 = nums[0], nums[1], nums[4], nums[7], nums[10]
                if all(v is not None for v in (current, previous_month, previous_year, year10, year15)):
                    prev_month = date(year if month > 1 else year - 1, month - 1 if month > 1 else 12, 1)
                    pensioner_points[terr] = [
                        (f"{year-15:04d}-{month:02d}", int(year15)),
                        (f"{year-10:04d}-{month:02d}", int(year10)),
                        (f"{year-1:04d}-{month:02d}", int(previous_year)),
                        (f"{prev_month.year:04d}-{prev_month.month:02d}", int(previous_month)),
                        (period, int(current)),
                    ]

    # La página histórica aporta una serie nacional mucho más larga.
    history_text = "\n".join((page.extract_text() or "") for page in reader.pages[-6:])
    history_points_pensioners: list[tuple[str, int]] = []
    history_points_pensions: list[tuple[str, int]] = []
    history_points_ratio: list[tuple[str, float]] = []
    current_year: int | None = None
    for line in history_text.splitlines():
        m = re.match(r"\s*(?:(20\d{2})\s+)?(Enero|Febrero|Marzo|Abril|Mayo|Junio|Julio|Agosto|Septiembre|Octubre|Noviembre|Diciembre)\s+(\d[\d.]*)\s+(\d[\d.]*)\s+(\d,\d+)", line, flags=re.I)
        if not m:
            continue
        if m.group(1): current_year = int(m.group(1))
        if current_year is None: continue
        month_no = month_names[normalize_key(m.group(2))]
        hist_period = f"{current_year:04d}-{month_no:02d}"
        history_points_pensioners.append((hist_period, int(parse_spanish_number(m.group(3)))))
        history_points_pensions.append((hist_period, int(parse_spanish_number(m.group(4)))))
        history_points_ratio.append((hist_period, round(parse_spanish_number(m.group(5)), 3)))
    if history_points_pensioners:
        pensioner_points[("ES", "España")] = history_points_pensioners
        # Sustituimos la observación nacional de número de pensiones por su serie histórica.
        for series in output:
            if series["id"] == "pensions_total":
                for item in series["territories"]:
                    if item["code"] == "ES": item["points"] = make_points(history_points_pensions)

    pensioners = series_base(
        id_="pensioners_total", name="Pensionistas del sistema de la Seguridad Social",
        description="Personas titulares de una o más pensiones del sistema; cada persona se contabiliza una sola vez.",
        unit="personas", frequency="Mensual", source="INSS · Evolución mensual de pensionistas",
        source_url=PENSIONS_PAGE,
    )
    pensioners["territories"] = order_territories([
        {"code": code, "name": name, "points": make_points(points)}
        for (code, name), points in pensioner_points.items()
    ])
    output.append(pensioners)

    ratio = series_base(
        id_="pensions_per_pensioner", name="Pensiones por pensionista",
        description="Promedio de pensiones contributivas en vigor por cada pensionista.",
        unit="pensiones por persona", frequency="Mensual", source="INSS · Evolución mensual de pensionistas",
        source_url=PENSIONS_PAGE, decimals=3,
    )
    ratio["territories"] = [{"code": "ES", "name": "España", "points": make_points(history_points_ratio)}]
    output.append(ratio)

    check = SourceCheck("pensions", True, datetime.now(MADRID).isoformat(timespec="seconds"),
                        f"{len(output)} series; corte {period}; {len(rows)} territorios",
                        sha256(content), fetched_from or PENSIONS_PAGE)
    return output, check


# --------------------------- Ingreso Mínimo Vital -------------------------

def discover_imv_xlsx() -> str:
    candidates: list[str] = []
    for page_url in (IMV_LATEST_PAGE, "https://www.inclusion.gob.es/web/policy-lab/inicio", IMV_STATS_PAGE):
        try:
            page = fetch_text(page_url)
            parser = LinkCollector(); parser.feed(page)
            for href, text in parser.links:
                joined = normalize_key(f"{href} {text}")
                if ".XLSX" in joined and ("COMPLETO" in joined or "CCAA" in joined or "COMUNIDAD" in joined):
                    candidates.append(urljoin(page_url, html.unescape(href)))
        except Exception:
            continue
    if not candidates:
        raise RuntimeError("No se localizó el dossier territorial del IMV")
    return max(candidates, key=lambda url: (re.findall(r"(20\d{4})", url) or ["000000"])[-1])

discover_imv_xlsx.page = IMV_STATS_PAGE

MONTHS_ES = {name: i for i, name in enumerate(
    ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"], 1)}


def parse_month_period(value: Any) -> str | None:
    text = normalize_key(value)
    m = re.search(r"(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)(?:\s+DE)?\s+(20\d{2})", text)
    if not m: return None
    return f"{int(m.group(2)):04d}-{MONTHS_ES[m.group(1)]:02d}"


def parse_imv_sheet(workbook: Any, sheet_name: str, *, multiplier: float = 1.0,
                    decimals: int = 0) -> list[dict[str, Any]]:
    ws = workbook[sheet_name]
    period_cols: list[tuple[int, str]] = []
    for col in range(3, 6):
        period = parse_month_period(ws.cell(13, col).value)
        if period: period_cols.append((col, period))
    values: dict[tuple[str, str], list[tuple[str, float | int]]] = {}
    for row in range(14, ws.max_row + 1):
        first, second = ws.cell(row, 1).value, ws.cell(row, 2).value
        if normalize_key(first) == "TOTAL":
            terr = ("ES", "España")
        elif normalize_key(second) == "TOTAL":
            terr = territory(first)
        else:
            continue
        if not terr: continue
        points: list[tuple[str, float | int]] = []
        for col, period in period_cols:
            raw = try_number(ws.cell(row, col).value)
            if raw is None: continue
            value = raw * multiplier
            value = int(round(value)) if decimals == 0 else round(value, decimals)
            points.append((period, value))
        if points: values[terr] = points
    return order_territories([
        {"code": code, "name": name, "points": make_points(points)}
        for (code, name), points in values.items()
    ])


def collect_imv(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    content, fetched_from = local_or_discover(local_dir, "imv_latest.xlsx", discover_imv_xlsx)
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    specs = [
        ("imv_households", "Hogares perceptores del Ingreso Mínimo Vital", "Número de prestaciones activas del IMV, equivalente al número de hogares perceptores.", "hogares", "IMV. 2.1. Ev. Prestaciones", 1, 0),
        ("imv_beneficiaries", "Personas beneficiarias del Ingreso Mínimo Vital", "Personas que forman parte de hogares con una prestación activa del IMV.", "personas", "IMV. 2.2. Ev. Beneficiarios", 1, 0),
        ("imv_beneficiaries_per_household", "Beneficiarios del IMV por hogar", "Promedio de personas beneficiarias por prestación activa del IMV.", "personas por hogar", "IMV. 2.3. Ev. Ratio", 1, 3),
        ("imv_monthly_payroll", "Nómina mensual del Ingreso Mínimo Vital", "Importe bruto mensual de la nómina del IMV.", "euros al mes", "IMV. 2.4. Ev. Importe bruto", 1_000_000, 2),
        ("imv_average_per_household", "Cuantía media mensual del IMV por hogar", "Cuantía mensual media de la prestación por hogar perceptor.", "euros al mes", "IMV. 2.5. Ev. Cuantía hogar", 1, 2),
        ("imv_average_per_beneficiary", "Cuantía media mensual del IMV por beneficiario", "Cuantía mensual media de la prestación por persona beneficiaria.", "euros al mes", "IMV. 2.6. Ev. Cuantía benef", 1, 2),
    ]
    output: list[dict[str, Any]] = []
    for id_, name, description, unit, sheet, multiplier, decimals in specs:
        series = series_base(id_=id_, name=name, description=description, unit=unit, frequency="Mensual",
                             source="INSS · Estadística del Ingreso Mínimo Vital", source_url=IMV_STATS_PAGE,
                             decimals=decimals)
        series["territories"] = parse_imv_sheet(wb, sheet, multiplier=multiplier, decimals=decimals)
        if len(series["territories"]) < 18:
            raise RuntimeError(f"{sheet}: solo {len(series['territories'])} territorios")
        output.append(series)
    latest = output[0]["territories"][0]["points"][-1]["period"]
    return output, SourceCheck("imv", True, datetime.now(MADRID).isoformat(timespec="seconds"),
                               f"{len(output)} series; corte {latest}; 20 territorios", sha256(content), fetched_from or IMV_STATS_PAGE)


# ----------------------------- Dependencia --------------------------------

def discover_imserso_xlsx() -> str:
    return discover_link(IMSERSO_MONTHLY_PAGE, lambda joined: "ESTSISAAD_" in normalize_key(joined) and ".XLSX" in normalize_key(joined), rank_pattern=r"estsisaad_(\d{8})")

discover_imserso_xlsx.page = IMSERSO_MONTHLY_PAGE


def normalize_excel_period(value: Any) -> str | None:
    if isinstance(value, datetime): return value.date().isoformat()
    if isinstance(value, date): return value.isoformat()
    text = str(value or "").strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y"):
        try: return datetime.strptime(text, fmt).date().isoformat()
        except ValueError: pass
    return None


def parse_saad_evolution(wb: Any, sheet_name: str) -> list[dict[str, Any]]:
    ws = wb[sheet_name]
    periods: list[tuple[int, str]] = []
    for col in range(4, min(ws.max_column, 20) + 1):
        period = normalize_excel_period(ws.cell(7, col).value)
        if period: periods.append((col, period))
    values: dict[tuple[str, str], list[tuple[str, int]]] = {}
    for row in range(9, ws.max_row + 1):
        terr = territory(ws.cell(row, 2).value)
        if not terr: continue
        points = [(period, int(ws.cell(row, col).value)) for col, period in periods if isinstance(ws.cell(row, col).value, (int, float))]
        if points: values[terr] = points
    return order_territories([{"code": c, "name": n, "points": make_points(p)} for (c, n), p in values.items()])


def parse_saad_snapshot(wb: Any, sheet_name: str, value_col: int, period: str) -> list[dict[str, Any]]:
    ws = wb[sheet_name]
    result = []
    for row in range(1, ws.max_row + 1):
        terr = territory(ws.cell(row, 2).value)
        raw = ws.cell(row, value_col).value
        if terr and isinstance(raw, (int, float)):
            result.append({"code": terr[0], "name": terr[1], "points": [{"period": period, "value": int(round(raw))}]})
    return order_territories(result)


def collect_dependency_extended(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    content, fetched_from = local_or_discover(local_dir, "estsisaad_20260630.xlsx", discover_imserso_xlsx)
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    latest_period = None
    for col in range(4, 20):
        p = normalize_excel_period(wb["EVO_sol"].cell(7, col).value)
        if p: latest_period = p
    if not latest_period: raise RuntimeError("No se identificó el periodo del SAAD")

    specs = [
        ("dependency_applications", "Solicitudes de dependencia", "Solicitudes en vigor registradas en el Sistema para la Autonomía y Atención a la Dependencia.", "solicitudes", "EVO_sol"),
        ("dependency_degree_resolutions", "Resoluciones de grado de dependencia", "Resoluciones de grado en vigor registradas en el SAAD.", "resoluciones", "EVO_resol"),
        ("dependency_entitled_people", "Personas con derecho a prestación por dependencia", "Personas con grado reconocido que genera derecho a prestaciones del SAAD.", "personas", "EVO_derecho"),
        ("dependency_pia_resolutions", "Resoluciones de Programa Individual de Atención (PIA)", "Personas con resolución de su Programa Individual de Atención.", "personas", "EVO_resolPIA"),
        ("dependency_waiting_pia", "Personas con derecho sin resolución de PIA", "Personas con derecho reconocido que todavía no disponen de resolución de PIA.", "personas", "EVO_sinPIA"),
        ("dependency_total_benefits", "Prestaciones de dependencia reconocidas", "Número total de prestaciones reconocidas en el SAAD; una persona puede tener más de una.", "prestaciones", "EVO_prest"),
    ]
    output: list[dict[str, Any]] = []
    for id_, name, description, unit, sheet in specs:
        series = series_base(id_=id_, name=name, description=description, unit=unit,
                             frequency="Mensual (cortes seleccionados)", source="Imserso · Estadística mensual del SAAD",
                             source_url=IMSERSO_MONTHLY_PAGE)
        series["territories"] = parse_saad_evolution(wb, sheet)
        output.append(series)

    snapshots = [
        ("dependency_wait_request_to_degree", "Tiempo de espera: solicitud a resolución de grado", "Tiempo medio desde la solicitud de dependencia hasta la resolución de grado.", "días", "9TiempoEspera", 5),
        ("dependency_wait_degree_to_benefit", "Tiempo de espera: grado a resolución de prestación", "Tiempo medio desde la resolución de grado hasta la resolución de prestación.", "días", "9TiempoEspera", 8),
        ("dependency_wait_request_to_benefit", "Tiempo total de espera en dependencia", "Tiempo medio desde la solicitud de dependencia hasta la resolución de prestación.", "días", "9TiempoEspera", 11),
        ("dependency_pending_degree", "Solicitantes pendientes de resolución de grado", "Personas solicitantes que todavía no tienen resolución de grado.", "personas", "10pendResol", 3),
        ("dependency_pending_degree_over_6_months", "Solicitantes pendientes de grado durante seis meses o más", "Personas solicitantes cuya resolución de grado lleva pendiente al menos seis meses.", "personas", "10pendResol", 8),
        ("dependency_waiting_pia_over_6_months", "Personas pendientes de PIA durante seis meses o más", "Personas con derecho reconocido cuya resolución de PIA lleva pendiente al menos seis meses.", "personas", "10pendPrest", 8),
        ("dependency_effective_beneficiaries", "Beneficiarios con prestación de dependencia efectiva", "Personas con resolución de PIA que ya reciben al menos una prestación efectiva.", "personas", "12BenefEfect", 6),
        ("dependency_pia_without_effective_benefit", "Personas con PIA aún sin prestación efectiva", "Personas con resolución de PIA que todavía no reciben ninguna prestación efectiva.", "personas", "12BenefEfect", 9),
    ]
    for id_, name, description, unit, sheet, col in snapshots:
        series = series_base(id_=id_, name=name, description=description, unit=unit, frequency="Mensual",
                             source="Imserso · Estadística mensual del SAAD", source_url=IMSERSO_MONTHLY_PAGE)
        series["territories"] = parse_saad_snapshot(wb, sheet, col, latest_period)
        output.append(series)
    if any(len(s["territories"]) < 18 for s in output):
        bad = [s["id"] for s in output if len(s["territories"]) < 18]
        raise RuntimeError(f"Series SAAD incompletas: {bad}")
    return output, SourceCheck("imserso_saad_extended", True, datetime.now(MADRID).isoformat(timespec="seconds"),
                               f"{len(output)} series; corte {latest_period}; 20 territorios", sha256(content), fetched_from or IMSERSO_MONTHLY_PAGE)


# -------------------------- Listas de espera -------------------------------

def discover_health_pdf() -> str:
    return discover_link(HEALTH_WAITING_PAGE, lambda joined: ".PDF" in normalize_key(joined) and "DATOS POR COMUNIDADES" in normalize_key(joined))

discover_health_pdf.page = HEALTH_WAITING_PAGE


def parse_pdf_table_segment(text: str, start_marker: str, end_marker: str | None = None) -> dict[tuple[str, str], list[float | None]]:
    # Se trabaja por líneas: los índices de una cadena normalizada no coinciden
    # necesariamente con los del original cuando hay tildes o espacios especiales.
    lines = text.splitlines()
    start_at = 0
    for i, line in enumerate(lines):
        if normalize_key(start_marker) in normalize_key(line):
            start_at = i + 1
            break
    end_at = len(lines)
    if end_marker:
        for i in range(start_at, len(lines)):
            if normalize_key(end_marker) in normalize_key(lines[i]):
                end_at = i
                break
    result: dict[tuple[str, str], list[float | None]] = {}
    for line in lines[start_at:end_at]:
        matched = match_line_territory(line)
        if not matched: continue
        terr, remainder = matched
        nums = extract_numbers(remainder, preserve_missing=True)
        if nums: result[terr] = nums
    return result


def weighted_total(values: dict[tuple[str, str], list[float | None]], index: int,
                   weights: dict[tuple[str, str], float]) -> float | None:
    pairs = [(row[index], weights.get(terr)) for terr, row in values.items() if terr[0] != "ES" and len(row) > index]
    pairs = [(float(v), float(w)) for v, w in pairs if v is not None and w not in (None, 0)]
    if not pairs: return None
    return sum(v * w for v, w in pairs) / sum(w for _, w in pairs)


def collect_health_waiting(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    content, fetched_from = local_or_discover(local_dir, "health_waiting_latest.pdf", discover_health_pdf)
    reader = PdfReader(io.BytesIO(content))
    texts = [page.extract_text() or "" for page in reader.pages]
    joined = "\n".join(texts[:2])
    match = re.search(r"31 DE (ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE) DE (20\d{2})", normalize_key(joined))
    if not match:
        match = re.search(r"31 DE DICIEMBRE DE (20\d{2})", normalize_key("\n".join(texts)))
        period = f"{match.group(1)}-12-31" if match else "2025-12-31"
    else:
        period = f"{match.group(2)}-{MONTHS_ES[match.group(1)]:02d}-31"

    # Localiza páginas por sus títulos; funciona tanto con la versión de 9 como con la de 20 páginas.
    surgical_counts_text = next(t for t in texts if "NUMERO DE PACIENTES EN ESPERA ESTRUCTURAL" in normalize_key(t) and "CIRUGIA GENERAL" in normalize_key(t))
    surgical_time_text = next(t for t in texts if "TIEMPO MEDIO DE ESPERA" in normalize_key(t) and "PORCENTAJE DE PACIENTES CON MAS DE SEIS MESES" in normalize_key(t))
    consultation_counts_text = next(t for t in texts if "NUMERO DE PACIENTES EN ESPERA ESTRUCTURAL" in normalize_key(t) and "NEUROLOGIA" in normalize_key(t))
    consultation_time_text = next(t for t in texts if "TIEMPO MEDIO DE ESPERA (DIAS)" in normalize_key(t) and "MAS DE 60 DIAS" in normalize_key(t))

    s_counts = parse_pdf_table_segment(surgical_counts_text, "Número de pacientes en espera estructural", "Tasas de pacientes")
    s_rates = parse_pdf_table_segment(surgical_counts_text, "Tasas de pacientes en espera por 1000 habitantes", "Fuente:")
    s_days = parse_pdf_table_segment(surgical_time_text, "Tiempo medio de espera", "Porcentaje de pacientes")
    s_over = parse_pdf_table_segment(surgical_time_text, "Porcentaje de pacientes con más de seis meses", "Fuente:")
    c_counts = parse_pdf_table_segment(consultation_counts_text, "Número de pacientes en espera estructural", "Pacientes por 1000")
    c_rates = parse_pdf_table_segment(consultation_counts_text, "Pacientes por 1000 habitantes", "Fuente:")
    c_days = parse_pdf_table_segment(consultation_time_text, "Tiempo medio de espera", "Porcentaje de pacientes")
    c_over = parse_pdf_table_segment(consultation_time_text, "Porcentaje de pacientes que esperan más de 60 días", "Fuente:")
    if min(map(len, (s_counts, s_rates, s_days, s_over, c_counts, c_rates, c_days, c_over))) < 18:
        raise RuntimeError("No se reconocieron todas las comunidades en las listas de espera")

    # Totales nacionales. Los recuentos se suman; tiempos y porcentajes se ponderan por pacientes.
    s_total_counts = {terr: row[14] for terr, row in s_counts.items() if len(row) > 14}
    c_total_counts = {terr: row[10] for terr, row in c_counts.items() if len(row) > 10}
    s_counts[("ES", "España")] = [sum(float(row[i] or 0) for terr, row in s_counts.items() if terr[0] != "ES" and len(row) > i) for i in range(15)]
    c_counts[("ES", "España")] = [sum(float(row[i] or 0) for terr, row in c_counts.items() if terr[0] != "ES" and len(row) > i) for i in range(11)]
    # Las tasas nacionales se calculan a partir de la población implícita en la tasa total autonómica.
    def national_rate(counts, rates, total_idx):
        populations = []
        total = 0.0
        for terr, row in counts.items():
            if terr[0] == "ES" or len(row) <= total_idx or terr not in rates or len(rates[terr]) <= total_idx: continue
            count, rate = row[total_idx], rates[terr][total_idx]
            if count is not None and rate not in (None, 0):
                total += float(count); populations.append(float(count) / float(rate) * 1000)
        return total / sum(populations) * 1000 if populations else None
    s_rates[("ES", "España")] = [None] * 14 + [national_rate(s_counts, s_rates, 14)]
    c_rates[("ES", "España")] = [None] * 10 + [national_rate(c_counts, c_rates, 10)]
    s_days[("ES", "España")] = [None] * 14 + [weighted_total(s_days, 14, s_total_counts)]
    s_over[("ES", "España")] = [None] * 14 + [weighted_total(s_over, 14, s_total_counts)]
    c_days[("ES", "España")] = [None] * 10 + [weighted_total(c_days, 10, c_total_counts)]
    c_over[("ES", "España")] = [None] * 10 + [weighted_total(c_over, 10, c_total_counts)]

    def metric_values(table, idx, decimals=0):
        out = {}
        for terr, row in table.items():
            if len(row) > idx and row[idx] is not None:
                value = float(row[idx]); out[terr] = int(round(value)) if decimals == 0 else round(value, decimals)
        return out

    specs = [
        ("surgical_waiting_patients", "Pacientes en lista de espera quirúrgica", "Pacientes en espera estructural para una intervención quirúrgica programada.", "personas", s_counts, 14, 0),
        ("surgical_waiting_rate", "Tasa de lista de espera quirúrgica", "Pacientes en espera quirúrgica por cada 1.000 personas con tarjeta sanitaria.", "pacientes por 1.000 habitantes", s_rates, 14, 2),
        ("surgical_waiting_days", "Tiempo medio de espera quirúrgica", "Promedio de días naturales de espera de los pacientes pendientes de intervención.", "días", s_days, 14, 1),
        ("surgical_waiting_over_6_months", "Pacientes quirúrgicos con más de seis meses de espera", "Porcentaje de pacientes de la lista quirúrgica que llevan más de seis meses esperando.", "%", s_over, 14, 2),
        ("surgical_general_digestive_waiting", "Lista quirúrgica: cirugía general y digestiva", "Pacientes en espera estructural de cirugía general y del aparato digestivo.", "personas", s_counts, 0, 0),
        ("surgical_ophthalmology_waiting", "Lista quirúrgica: oftalmología", "Pacientes en espera estructural de cirugía oftalmológica.", "personas", s_counts, 2, 0),
        ("surgical_traumatology_waiting", "Lista quirúrgica: traumatología", "Pacientes en espera estructural de cirugía traumatológica.", "personas", s_counts, 4, 0),
        ("consultation_waiting_patients", "Pacientes pendientes de primera consulta especializada", "Pacientes en espera estructural de una primera consulta en atención especializada.", "personas", c_counts, 10, 0),
        ("consultation_waiting_rate", "Tasa de espera para primera consulta especializada", "Pacientes pendientes de primera consulta por cada 1.000 personas con tarjeta sanitaria.", "pacientes por 1.000 habitantes", c_rates, 10, 2),
        ("consultation_waiting_days", "Tiempo medio de espera para primera consulta", "Promedio de días naturales de espera para una primera consulta especializada.", "días", c_days, 10, 1),
        ("consultation_waiting_over_60_days", "Primeras consultas con más de 60 días de espera", "Porcentaje de pacientes cuya primera consulta supera los 60 días de demora.", "%", c_over, 10, 2),
        ("consultation_ophthalmology_waiting", "Primera consulta: oftalmología", "Pacientes pendientes de una primera consulta de oftalmología.", "personas", c_counts, 1, 0),
        ("consultation_traumatology_waiting", "Primera consulta: traumatología", "Pacientes pendientes de una primera consulta de traumatología.", "personas", c_counts, 2, 0),
        ("consultation_dermatology_waiting", "Primera consulta: dermatología", "Pacientes pendientes de una primera consulta de dermatología.", "personas", c_counts, 3, 0),
    ]
    output = [single_period_series(
        id_=id_, name=name, description=description, unit=unit, frequency="Semestral",
        source="Ministerio de Sanidad · SISLE-SNS", source_url=HEALTH_WAITING_PAGE,
        period=period, values=metric_values(table, idx, decimals), decimals=decimals,
        status="Oficial · datos suministrados por las comunidades autónomas",
    ) for id_, name, description, unit, table, idx, decimals in specs]
    return output, SourceCheck("health_waiting", True, datetime.now(MADRID).isoformat(timespec="seconds"),
                               f"{len(output)} series; corte {period}; 20 territorios en indicadores totales", sha256(content), fetched_from or HEALTH_WAITING_PAGE)


# ------------------------- Residencias de mayores -------------------------

def discover_residences_pdf() -> str:
    # La página anual contiene el informe PDF y sus anexos.
    return discover_link(RESIDENCES_PAGE, lambda joined: ".PDF" in normalize_key(joined) and ("SSPPMM" in normalize_key(joined) or "SERVICIOS SOCIALES" in normalize_key(joined)))

discover_residences_pdf.page = RESIDENCES_PAGE


def parse_residential_page(text: str) -> dict[tuple[str, str], list[float | None]]:
    result = {}
    for line in text.splitlines():
        matched = match_line_territory(line)
        if matched:
            terr, remainder = matched
            nums = extract_numbers(remainder, preserve_missing=True)
            if nums: result[terr] = nums
    return result


def collect_residences(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    content, fetched_from = local_or_discover(local_dir, "residences_latest.pdf", discover_residences_pdf)
    reader = PdfReader(io.BytesIO(content))
    texts = [page.extract_text() or "" for page in reader.pages]
    def best_page(*markers: str) -> tuple[str, dict[tuple[str, str], list[float | None]]]:
        candidates = [t for t in texts if all(normalize_key(m) in normalize_key(t) for m in markers)]
        if not candidates:
            raise RuntimeError(f"No se encontró la tabla residencial: {markers}")
        parsed = [(t, parse_residential_page(t)) for t in candidates]
        return max(parsed, key=lambda pair: len(pair[1]))
    centers_text, centers = best_page("CENTROS Y SU DISTRIBUCION POR TITULARIDAD")
    places_text, places = best_page("PLAZAS DE CENTROS RESIDENCIALES", "TIPO DE FINANCIACION")
    users_text, users = best_page("PERSONAS USUARIAS DE CENTROS RESIDENCIALES")
    prices_text, prices = best_page("PRECIO DEL SERVICIO DE CENTROS RESIDENCIALES")
    if min(map(len, (centers, places, users, prices))) < 18:
        raise RuntimeError("No se reconocieron todos los territorios del informe residencial")

    def vals(table, idx, decimals=0):
        result = {}
        for terr, row in table.items():
            if len(row) > idx and row[idx] is not None:
                v = float(row[idx]); result[terr] = int(round(v)) if decimals == 0 else round(v, decimals)
        return result
    specs = [
        ("residential_centers_total", "Centros residenciales para personas mayores", "Número total de centros residenciales para personas mayores.", "centros", centers, 3, 0),
        ("residential_centers_public", "Centros residenciales públicos", "Centros residenciales de titularidad pública.", "centros", centers, 4, 0),
        ("residential_centers_private", "Centros residenciales privados", "Centros residenciales de titularidad privada.", "centros", centers, 5, 0),
        ("residential_places_total", "Plazas en centros residenciales", "Número total de plazas en centros residenciales para personas mayores.", "plazas", places, 1, 0),
        ("residential_places_public_financing", "Plazas residenciales de financiación pública", "Plazas de centros residenciales financiadas públicamente, incluidas las concertadas.", "plazas", places, 2, 0),
        ("residential_places_private_financing", "Plazas residenciales de financiación privada", "Plazas de centros residenciales financiadas privadamente.", "plazas", places, 4, 0),
        ("residential_coverage", "Cobertura de plazas residenciales", "Plazas en centros residenciales por cada 100 personas de 65 o más años.", "plazas por 100 personas de 65+", places, 6, 2),
        ("residential_users", "Personas usuarias de centros residenciales", "Personas usuarias de centros residenciales para mayores.", "personas", users, 0, 0),
        ("residential_public_price", "Precio público anual de una plaza residencial", "Precio público anual medio por persona usuaria de centro residencial.", "euros al año", prices, 0, 2),
        ("residential_public_user_contribution", "Aportación del usuario a una plaza residencial pública", "Aportación anual media de la persona usuaria al precio público residencial.", "euros al año", prices, 1, 2),
        ("residential_concerted_price", "Precio anual de concertación de una plaza residencial", "Precio anual medio de concertación de una plaza residencial.", "euros al año", prices, 3, 2),
        ("residential_concerted_user_contribution", "Aportación del usuario a una plaza residencial concertada", "Aportación anual media de la persona usuaria al precio de una plaza concertada.", "euros al año", prices, 4, 2),
    ]
    output = [single_period_series(
        id_=id_, name=name, description=description, unit=unit, frequency="Anual",
        source="Imserso · Servicios sociales dirigidos a personas mayores", source_url=RESIDENCES_PAGE,
        period="2024", values=vals(table, idx, decimals), decimals=decimals,
    ) for id_, name, description, unit, table, idx, decimals in specs]
    return output, SourceCheck("residences", True, datetime.now(MADRID).isoformat(timespec="seconds"),
                               f"{len(output)} series; año 2024; 20 territorios", sha256(content), fetched_from or RESIDENCES_PAGE)


# ----------------------- Series educativas derivadas ----------------------
EDUCATION_STAGES = [
    ("general", "Enseñanzas de régimen general"), ("infant", "Educación Infantil"),
    ("early_childhood", "Primer ciclo de Educación Infantil (0-3)"), ("primary", "Educación Primaria"),
    ("eso", "Educación Secundaria Obligatoria"), ("bachiller", "Bachillerato"),
    ("fp_basic", "FP de Grado Básico"), ("fp_middle", "FP de Grado Medio"),
    ("fp_higher", "FP de Grado Superior"),
]


def derive_education_series(series_by_id: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for key, label in EDUCATION_STAGES:
        public = series_by_id.get(f"education_{key}_public")
        private = series_by_id.get(f"education_{key}_private")
        if not public or not private: continue
        pub_map = {(t["code"], p["period"]): p["value"] for t in public["territories"] for p in t["points"]}
        pri_map = {(t["code"], p["period"]): p["value"] for t in private["territories"] for p in t["points"]}
        names = {t["code"]: t["name"] for t in public["territories"] + private["territories"]}
        totals: dict[str, list[tuple[str, int]]] = {}; pubshares: dict[str, list[tuple[str, float]]] = {}; prishares: dict[str, list[tuple[str, float]]] = {}
        for code, period in sorted(set(pub_map) & set(pri_map)):
            total = pub_map[(code, period)] + pri_map[(code, period)]
            if total <= 0: continue
            totals.setdefault(code, []).append((period, int(total)))
            pubshares.setdefault(code, []).append((period, round(pub_map[(code, period)] / total * 100, 2)))
            prishares.setdefault(code, []).append((period, round(pri_map[(code, period)] / total * 100, 2)))
        source_url = public["source_url"]
        specs = [
            (f"education_{key}_total", f"{label}: alumnado total", f"Alumnado total de {label.lower()}, suma de centros públicos y privados.", "alumnos", totals, 0),
            (f"education_{key}_public_share", f"{label}: porcentaje del alumnado en centros públicos", f"Porcentaje del alumnado de {label.lower()} matriculado en centros públicos.", "%", pubshares, 2),
            (f"education_{key}_private_share", f"{label}: porcentaje del alumnado en centros privados", f"Porcentaje del alumnado de {label.lower()} matriculado en centros privados, incluidos concertados y no concertados.", "%", prishares, 2),
        ]
        for id_, name, description, unit, values, decimals in specs:
            series = series_base(id_=id_, name=name, description=description, unit=unit,
                                 frequency="Curso académico", source="Cálculo propio a partir de EDUCAbase",
                                 source_url=source_url, status="Calculado a partir de datos oficiales", decimals=decimals)
            series["territories"] = order_territories([
                {"code": code, "name": names.get(code, code), "points": make_points(points)}
                for code, points in values.items()
            ])
            output.append(series)
    return output


# ---------------- Catálogo de Atención Primaria -------------------------

PRIMARY_CARE_TERRITORIES: dict[str, tuple[str, str]] = {
    "SISTEMA NACIONAL DE SALUD": ("ES", "España"),
    "ANDALUCIA": ("ES-AN", "Andalucía"), "ARAGON": ("ES-AR", "Aragón"),
    "PRINCIPADO DE ASTURIAS": ("ES-AS", "Asturias"),
    "ILLES BALEARS": ("ES-IB", "Illes Balears"), "CANARIAS": ("ES-CN", "Canarias"),
    "CANTABRIA": ("ES-CB", "Cantabria"), "CASTILLA Y LEON": ("ES-CL", "Castilla y León"),
    "CASTILLA - LA MANCHA": ("ES-CM", "Castilla-La Mancha"),
    "CATALUNA": ("ES-CT", "Cataluña"), "COMUNIDAD VALENCIANA": ("ES-VC", "Comunitat Valenciana"),
    "EXTREMADURA": ("ES-EX", "Extremadura"), "GALICIA": ("ES-GA", "Galicia"),
    "COMUNIDAD DE MADRID": ("ES-MD", "Comunidad de Madrid"),
    "REGION DE MURCIA": ("ES-MC", "Región de Murcia"),
    "COMUNIDAD FORAL DE NAVARRA": ("ES-NC", "Comunidad Foral de Navarra"),
    "PAIS VASCO": ("ES-PV", "País Vasco"), "LA RIOJA": ("ES-RI", "La Rioja"),
    "CEUTA Y MELILLA": ("ES-CMEL", "Ceuta y Melilla"),
}


def primary_care_territory(value: Any) -> tuple[str, str] | None:
    key = normalize_key(value).replace("–", "-").replace("—", "-")
    key = re.sub(r"\s+", " ", key).strip()
    return PRIMARY_CARE_TERRITORIES.get(key)


def collect_primary_care_centers(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    """Centros, organización y modalidades de gestión de Atención Primaria."""
    filename = "2026_C_Catal_Centros_AP.xlsx"
    if local_dir and (local_dir / filename).exists():
        content = (local_dir / filename).read_bytes()
    else:
        content = fetch_bytes(PRIMARY_CARE_CATALOG_XLSX, timeout=180)
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    period = "2025-12-31"

    values: dict[str, dict[tuple[str, str], float | int | None]] = {
        key: {} for key in [
            "areas", "basic_zones", "centers_total", "health_centers", "local_consultories",
            "teaching_centers", "centers_public_direct", "centers_public_other", "centers_private",
            "health_public_direct", "health_public_other", "health_private",
            "consultory_public_direct", "consultory_public_other", "consultory_private",
        ]
    }
    general = wb["Datos generales"]
    for row in general.iter_rows(min_row=4, max_row=22, values_only=True):
        terr = primary_care_territory(row[0])
        if not terr:
            continue
        area_match = re.search(r"(\d+)", str(row[1] or ""))
        values["areas"][terr] = int(area_match.group(1)) if area_match else None
        values["basic_zones"][terr] = int(try_number(row[2]) or 0)
        values["centers_total"][terr] = int(try_number(row[3]) or 0)
        values["health_centers"][terr] = int(try_number(row[4]) or 0)
        values["local_consultories"][terr] = int(try_number(row[5]) or 0)

    stats = wb["Otras estadísticas"]
    for row in stats.iter_rows(min_row=4, max_row=22, values_only=True):
        terr = primary_care_territory(row[0]) or primary_care_territory(row[11])
        if not terr:
            continue
        values["teaching_centers"][terr] = int(try_number(row[1]) or 0)
        # Columnas 13–18 de la hoja: tres modalidades para centro de salud y consultorio.
        hp, ho, hv = (int(try_number(row[i]) or 0) for i in (12, 13, 14))
        cp, co, cv = (int(try_number(row[i]) or 0) for i in (15, 16, 17))
        values["health_public_direct"][terr] = hp
        values["health_public_other"][terr] = ho
        values["health_private"][terr] = hv
        values["consultory_public_direct"][terr] = cp
        values["consultory_public_other"][terr] = co
        values["consultory_private"][terr] = cv
        values["centers_public_direct"][terr] = hp + cp
        values["centers_public_other"][terr] = ho + co
        values["centers_private"][terr] = hv + cv

    specs = [
        ("primary_care_health_areas", "Áreas de salud", "Áreas o estructuras territoriales equivalentes de Atención Primaria.", "áreas", "areas"),
        ("primary_care_basic_zones", "Zonas básicas de salud", "Zonas básicas de salud del Sistema Nacional de Salud.", "zonas", "basic_zones"),
        ("primary_care_centers_total", "Centros de Atención Primaria", "Total de centros de salud y consultorios locales del SNS.", "centros", "centers_total"),
        ("primary_care_health_centers", "Centros de salud", "Centros de salud del Sistema Nacional de Salud.", "centros", "health_centers"),
        ("primary_care_local_consultories", "Consultorios locales", "Consultorios locales de Atención Primaria del SNS.", "consultorios", "local_consultories"),
        ("primary_care_teaching_centers", "Centros de Atención Primaria con acreditación docente", "Centros cabecera acreditados para formación sanitaria especializada.", "centros", "teaching_centers"),
        ("primary_care_centers_public_direct", "Centros de Atención Primaria de gestión pública directa", "Centros cuya producción asistencial corresponde directamente al servicio público de salud.", "centros", "centers_public_direct"),
        ("primary_care_centers_public_other", "Centros de Atención Primaria con otras modalidades públicas", "Centros gestionados por empresas, fundaciones públicas o consorcios.", "centros", "centers_public_other"),
        ("primary_care_centers_private", "Centros de Atención Primaria de gestión privada", "Centros del SNS cuya producción del servicio se clasifica como privada.", "centros", "centers_private"),
        ("primary_care_health_public_direct", "Centros de salud de gestión pública directa", "Centros de salud de gestión pública directa.", "centros", "health_public_direct"),
        ("primary_care_health_public_other", "Centros de salud con otras modalidades públicas", "Centros de salud gestionados por empresas, fundaciones públicas o consorcios.", "centros", "health_public_other"),
        ("primary_care_health_private", "Centros de salud de gestión privada", "Centros de salud del SNS clasificados como de gestión privada.", "centros", "health_private"),
        ("primary_care_consultory_public_direct", "Consultorios de gestión pública directa", "Consultorios locales de gestión pública directa.", "consultorios", "consultory_public_direct"),
        ("primary_care_consultory_public_other", "Consultorios con otras modalidades públicas", "Consultorios gestionados mediante otras modalidades públicas.", "consultorios", "consultory_public_other"),
        ("primary_care_consultory_private", "Consultorios de gestión privada", "Consultorios locales del SNS clasificados como de gestión privada.", "consultorios", "consultory_private"),
    ]
    output = [single_period_series(
        id_=series_id, name=name, description=description, unit=unit, frequency="Anual",
        source="Ministerio de Sanidad · Catálogo de Centros de Atención Primaria 2026",
        source_url=PRIMARY_CARE_CATALOG_PAGE, period=period, values=values[key], status="Oficial",
    ) for series_id, name, description, unit, key in specs]

    if len(output) != 15 or any(len(series["territories"]) < 18 for series in output):
        raise RuntimeError("Cobertura insuficiente en el Catálogo de Atención Primaria")
    check = SourceCheck(
        "primary_care_catalog", True, datetime.now(MADRID).isoformat(timespec="seconds"),
        f"{len(output)} series; situación {period}; {len(output[0]['territories'])} ámbitos territoriales",
        sha256(content), PRIMARY_CARE_CATALOG_PAGE,
    )
    return output, check


# ------------------- Catálogo Nacional de Hospitales --------------------

HOSPITAL_CCAA_CODES: dict[str, tuple[str, str]] = {
    "01": ("ES-AN", "Andalucía"), "02": ("ES-AR", "Aragón"),
    "03": ("ES-AS", "Asturias"), "04": ("ES-IB", "Illes Balears"),
    "05": ("ES-CN", "Canarias"), "06": ("ES-CB", "Cantabria"),
    "07": ("ES-CL", "Castilla y León"), "08": ("ES-CM", "Castilla-La Mancha"),
    "09": ("ES-CT", "Cataluña"), "10": ("ES-VC", "Comunitat Valenciana"),
    "11": ("ES-EX", "Extremadura"), "12": ("ES-GA", "Galicia"),
    "13": ("ES-MD", "Comunidad de Madrid"), "14": ("ES-MC", "Región de Murcia"),
    "15": ("ES-NC", "Comunidad Foral de Navarra"), "16": ("ES-PV", "País Vasco"),
    "17": ("ES-RI", "La Rioja"), "18": ("ES-CE", "Ceuta"),
    "19": ("ES-ML", "Melilla"),
}


def _hospital_values_with_spain(values: dict[tuple[str, str], float | int]) -> dict[tuple[str, str], float | int]:
    output = dict(values)
    output[("ES", "España")] = sum(values.values())
    return output


def collect_hospital_resources(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    """Hospitales, camas, altas/bajas y equipamiento del CNH 2025.

    El catálogo 2025 representa la situación de los centros a 31 de diciembre de
    2024. La dependencia funcional 1–8 se agrupa como pública y 20–22 como privada
    o no pública, siguiendo la codificación del propio catálogo.
    """
    if local_dir and (local_dir / "CNH_2025.xlsx").exists():
        content = (local_dir / "CNH_2025.xlsx").read_bytes()
    else:
        content = fetch_bytes(HOSPITAL_CATALOG_XLSX, timeout=180)
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    period = "2024-12-31"

    counts: dict[str, dict[tuple[str, str], float | int]] = {
        key: {} for key in [
            "centers_total", "centers_public", "centers_private",
            "beds_total", "beds_public", "beds_private",
            "class_general", "class_specialized", "class_long_stay",
            "class_mental_health", "class_other",
        ]
    }
    directory = wb["DIRECTORIO DE HOSPITALES"]
    directory_by_code: dict[str, tuple[tuple[str, str], bool]] = {}
    class_map = {
        "C11": "class_general", "C12": "class_specialized", "C13": "class_long_stay",
        "C14": "class_mental_health", "C190": "class_other",
    }
    for row in directory.iter_rows(min_row=2, values_only=True):
        ccaa_code = str(row[9] or "").zfill(2)
        terr = HOSPITAL_CCAA_CODES.get(ccaa_code)
        if not terr:
            continue
        dep_code_text = str(row[15] or "").strip()
        try:
            dep_code = int(float(dep_code_text))
        except ValueError:
            dep_code = 20
        is_public = dep_code < 20
        codcnh = str(row[1] or "").strip()
        if codcnh:
            directory_by_code[codcnh] = (terr, is_public)
        beds = try_number(row[12]) or 0
        for key, value in (("centers_total", 1), ("beds_total", beds)):
            counts[key][terr] = counts[key].get(terr, 0) + value
        owner = "public" if is_public else "private"
        counts[f"centers_{owner}"][terr] = counts[f"centers_{owner}"].get(terr, 0) + 1
        counts[f"beds_{owner}"][terr] = counts[f"beds_{owner}"].get(terr, 0) + beds
        class_key = class_map.get(str(row[13] or "").strip())
        if class_key:
            counts[class_key][terr] = counts[class_key].get(terr, 0) + 1

    # Equipamiento instalado. Las filas DC son componentes desagregados de un complejo
    # y no contienen cifras agregables; se omiten de manera explícita.
    equipment_specs = {
        15: ("hospital_ct_units", "Equipos de tomografía computarizada hospitalaria", "equipos TAC"),
        16: ("hospital_mri_units", "Equipos de resonancia magnética hospitalaria", "equipos RMN"),
        24: ("hospital_pet_units", "Equipos PET hospitalarios", "equipos PET"),
        25: ("hospital_mammography_units", "Mamógrafos hospitalarios", "mamógrafos"),
        27: ("hospital_dialysis_units", "Puestos de hemodiálisis hospitalaria", "puestos"),
    }
    equipment: dict[str, dict[tuple[str, str], float | int]] = {spec[0]: {} for spec in equipment_specs.values()}
    structure = wb["ESTRUCTURA FUNCIONAL"]
    for row in structure.iter_rows(min_row=2, values_only=True):
        terr = HOSPITAL_CCAA_CODES.get(str(row[5] or "").zfill(2))
        if not terr:
            continue
        for col_idx, (series_id, _, _) in equipment_specs.items():
            raw = row[col_idx]
            if str(raw).strip().upper() == "DC":
                continue
            value = try_number(raw)
            if value is not None:
                equipment[series_id][terr] = equipment[series_id].get(terr, 0) + value

    # Altas incorporadas al catálogo 2025, con titularidad recuperada del directorio.
    openings: dict[str, dict[tuple[str, str], int]] = {"total": {}, "public": {}, "private": {}}
    for row in wb["ALTAS DE HOSPITALES"].iter_rows(min_row=2, values_only=True):
        codcnh = str(row[1] or "").strip()
        info = directory_by_code.get(codcnh)
        if info:
            terr, is_public = info
        else:
            terr = territory(row[5])
            is_public = False
        if not terr:
            continue
        openings["total"][terr] = openings["total"].get(terr, 0) + 1
        owner = "public" if is_public else "private"
        openings[owner][terr] = openings[owner].get(terr, 0) + 1

    closures: dict[tuple[str, str], int] = {}
    for row in wb["BAJAS DE HOSPITALES"].iter_rows(min_row=2, values_only=True):
        terr = territory(row[4])
        if terr:
            closures[terr] = closures.get(terr, 0) + 1

    output: list[dict[str, Any]] = []
    count_specs = [
        ("hospital_centers_total", "Hospitales", "Centros incluidos en el Catálogo Nacional de Hospitales.", "centers_total", "hospitales"),
        ("hospital_centers_public", "Hospitales de dependencia funcional pública", "Hospitales dependientes de administraciones u organismos públicos.", "centers_public", "hospitales"),
        ("hospital_centers_private", "Hospitales de dependencia funcional privada", "Hospitales privados, de mutuas u organizaciones no gubernamentales según la clasificación funcional del catálogo.", "centers_private", "hospitales"),
        ("hospital_beds_total", "Camas hospitalarias instaladas", "Camas instaladas en los hospitales del catálogo.", "beds_total", "camas"),
        ("hospital_beds_public", "Camas en hospitales públicos", "Camas instaladas en hospitales de dependencia funcional pública.", "beds_public", "camas"),
        ("hospital_beds_private", "Camas en hospitales privados", "Camas instaladas en hospitales de dependencia funcional privada o no pública.", "beds_private", "camas"),
        ("hospital_general_centers", "Hospitales generales", "Centros clasificados como hospitales generales.", "class_general", "hospitales"),
        ("hospital_specialized_centers", "Hospitales especializados", "Centros clasificados como hospitales especializados.", "class_specialized", "hospitales"),
        ("hospital_long_stay_centers", "Hospitales de media y larga estancia", "Centros de media y larga estancia.", "class_long_stay", "hospitales"),
        ("hospital_mental_health_centers", "Hospitales de salud mental y toxicomanías", "Centros hospitalarios de salud mental y tratamiento de toxicomanías.", "class_mental_health", "hospitales"),
        ("hospital_other_inpatient_centers", "Otros centros con internamiento", "Otros centros con internamiento incluidos en el Catálogo Nacional de Hospitales.", "class_other", "hospitales"),
    ]
    for series_id, name, description, key, unit in count_specs:
        output.append(single_period_series(
            id_=series_id, name=name, description=description, unit=unit, frequency="Anual",
            source="Ministerio de Sanidad · Catálogo Nacional de Hospitales 2025",
            source_url=HOSPITAL_CATALOG_PAGE, period=period,
            values=_hospital_values_with_spain(counts[key]), status="Oficial",
        ))
    for series_id, name, unit in equipment_specs.values():
        output.append(single_period_series(
            id_=series_id, name=name,
            description=f"Número de {unit} instalados en los centros del Catálogo Nacional de Hospitales.",
            unit=unit, frequency="Anual", source="Ministerio de Sanidad · Catálogo Nacional de Hospitales 2025",
            source_url=HOSPITAL_CATALOG_PAGE, period=period,
            values=_hospital_values_with_spain(equipment[series_id]), status="Oficial",
        ))
    for owner, label in (("total", "Altas de hospitales en el catálogo"),
                         ("public", "Altas de hospitales públicos en el catálogo"),
                         ("private", "Altas de hospitales privados en el catálogo")):
        output.append(single_period_series(
            id_=f"hospital_openings_{owner}", name=label,
            description="Centros incorporados como altas en la edición 2025 del Catálogo Nacional de Hospitales.",
            unit="hospitales", frequency="Anual", source="Ministerio de Sanidad · Catálogo Nacional de Hospitales 2025",
            source_url=HOSPITAL_CATALOG_PAGE, period="2025",
            values=_hospital_values_with_spain(openings[owner]), status="Oficial",
        ))
    output.append(single_period_series(
        id_="hospital_closures_total", name="Bajas de hospitales en el catálogo",
        description="Centros dados de baja en la edición 2025 del Catálogo Nacional de Hospitales.",
        unit="hospitales", frequency="Anual", source="Ministerio de Sanidad · Catálogo Nacional de Hospitales 2025",
        source_url=HOSPITAL_CATALOG_PAGE, period="2025",
        values=_hospital_values_with_spain(closures), status="Oficial",
    ))

    if len(output) != 20:
        raise RuntimeError(f"Se esperaban 20 series hospitalarias y se generaron {len(output)}")
    check = SourceCheck(
        "hospital_catalog", True, datetime.now(MADRID).isoformat(timespec="seconds"),
        f"{len(output)} series; situación {period}; {len(directory_by_code)} hospitales identificados",
        sha256(content), HOSPITAL_CATALOG_PAGE,
    )
    return output, check


# -------------------- Vivienda: valor tasado por m² ----------------------

ISTAC_TERRITORIES: dict[str, tuple[str, str]] = {
    "ES": ("ES", "España"),
    "ES61": ("ES-AN", "Andalucía"), "ES24": ("ES-AR", "Aragón"),
    "ES12": ("ES-AS", "Asturias"), "ES53": ("ES-IB", "Illes Balears"),
    "ES70": ("ES-CN", "Canarias"), "ES13": ("ES-CB", "Cantabria"),
    "ES41": ("ES-CL", "Castilla y León"), "ES42": ("ES-CM", "Castilla-La Mancha"),
    "ES51": ("ES-CT", "Cataluña"), "ES52": ("ES-VC", "Comunitat Valenciana"),
    "ES43": ("ES-EX", "Extremadura"), "ES11": ("ES-GA", "Galicia"),
    "ES30": ("ES-MD", "Comunidad de Madrid"), "ES62": ("ES-MC", "Región de Murcia"),
    "ES22": ("ES-NC", "Comunidad Foral de Navarra"), "ES21": ("ES-PV", "País Vasco"),
    "ES23": ("ES-RI", "La Rioja"), "ES630": ("ES-CE", "Ceuta"),
    "ES640": ("ES-ML", "Melilla"),
}


def collect_house_prices(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    """Valor tasado medio de la vivienda y sus variaciones trimestral/anual.

    La API estadística de ISTAC reproduce la operación oficial del Ministerio de
    Vivienda y ofrece un CSV estable, apto para actualización automatizada.
    """
    if local_dir and (local_dir / "house_price_m2.csv").exists():
        content = (local_dir / "house_price_m2.csv").read_bytes()
    else:
        content = fetch_bytes(HOUSE_PRICE_CSV, timeout=240)

    specs = {
        "VIVIENDA__VALOR_TASADO": (
            "house_appraised_value_m2", "Valor tasado medio de la vivienda",
            "Valor tasado medio de la vivienda libre y protegida, en euros por metro cuadrado.",
            "euros por m²", 1,
        ),
        "VIVIENDA__VALOR_TASADO__TASA_VARIACION_ANUAL": (
            "house_appraised_value_annual_change", "Variación anual del valor tasado de la vivienda",
            "Tasa de variación interanual del valor tasado medio de la vivienda.", "%", 2,
        ),
        "VIVIENDA__VALOR_TASADO__TASA_VARIACION_INTERPERIODICA": (
            "house_appraised_value_quarterly_change", "Variación trimestral del valor tasado de la vivienda",
            "Tasa de variación respecto al trimestre anterior del valor tasado medio de la vivienda.", "%", 2,
        ),
    }
    observations: dict[str, dict[str, list[tuple[str, float]]]] = {
        spec[0]: {} for spec in specs.values()
    }
    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig", errors="replace")))
    for row in reader:
        if row.get("ANTIGUEDAD_CODE") != "_T":
            continue
        measure = row.get("MEDIDAS_CODE")
        if measure not in specs:
            continue
        terr = ISTAC_TERRITORIES.get(row.get("TERRITORIO_CODE", ""))
        period = row.get("TIME_PERIOD_CODE")
        raw = row.get("OBS_VALUE")
        if not terr or not period or raw in (None, ""):
            continue
        try:
            value = float(str(raw).replace(",", "."))
        except ValueError:
            continue
        series_id = specs[measure][0]
        observations[series_id].setdefault(terr[0], []).append((period, value))

    output: list[dict[str, Any]] = []
    latest = ""
    for measure, (series_id, name, description, unit, decimals) in specs.items():
        series = series_base(
            id_=series_id, name=name, description=description, unit=unit,
            frequency="Trimestral", source="Ministerio de Vivienda / ISTAC",
            source_url=HOUSE_PRICE_PAGE, status="Oficial", decimals=decimals,
        )
        territories_out = []
        for code, pairs in observations[series_id].items():
            terr = next((item for item in ISTAC_TERRITORIES.values() if item[0] == code), None)
            if not terr:
                continue
            points = make_points(pairs)
            if points:
                latest = max(latest, points[-1]["period"])
                territories_out.append({"code": code, "name": terr[1], "points": points})
        series["territories"] = order_territories(territories_out)
        if len(series["territories"]) < 18:
            raise RuntimeError(f"Cobertura territorial insuficiente en {series_id}")
        output.append(series)

    check = SourceCheck(
        "mivau_house_price", True, datetime.now(MADRID).isoformat(timespec="seconds"),
        f"{len(output)} series; último trimestre {latest or '—'}; 20 territorios",
        sha256(content), HOUSE_PRICE_PAGE,
    )
    return output, check


# ---------------------- Fondo de Reserva ---------------------------------

def collect_pension_reserve(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    """Serie nacional del Fondo de Reserva de la Seguridad Social.

    Los cierres anuales proceden de los informes oficiales; los dos últimos
    puntos son fotografías oficiales de noviembre de 2025 y marzo de 2026.
    """
    values: list[tuple[str, float]] = [
        ("2000", 604), ("2001", 2433), ("2002", 6169), ("2003", 12025),
        ("2004", 19330), ("2005", 27185), ("2006", 35879), ("2007", 45716),
        ("2008", 57223), ("2009", 60022), ("2010", 64375), ("2011", 66815),
        ("2012", 63008), ("2013", 53744), ("2014", 41634), ("2015", 32481),
        ("2016", 15020), ("2017", 8095), ("2018", 5043), ("2019", 2153),
        ("2020", 2138), ("2021", 2138), ("2022", 2141), ("2023", 5578.44),
        ("2024", 9376.70), ("2025-11", 13683.81), ("2026-03", 15267.00),
    ]
    series = series_base(
        id_="pension_reserve_fund", name="Fondo de Reserva de la Seguridad Social",
        description=("Saldo del Fondo de Reserva a precio de adquisición. Hasta 2024 se muestran "
                     "cierres anuales; 2025 y 2026 son cortes oficiales de noviembre y marzo."),
        unit="millones de euros", frequency="Anual / actualización extraordinaria",
        source="Seguridad Social · Informes del Fondo de Reserva",
        source_url=PENSION_RESERVE_URL, status="Oficial", decimals=2,
    )
    series["territories"] = [{"code": "ES", "name": "España", "points": make_points(values)}]
    encoded = repr(values).encode("utf-8")
    check = SourceCheck(
        "pension_reserve", True, datetime.now(MADRID).isoformat(timespec="seconds"),
        "1 serie; último corte 2026-03; ámbito nacional", sha256(encoded), PENSION_RESERVE_URL,
    )
    return [series], check


# --------------------- Sistema universitario -----------------------------

UNIVERSITY_REGIONAL_2024_25: dict[str, tuple[str, float, int, int, int, int]] = {
    "ES-AN": ("Andalucía", 27.3, 253326, 209445, 27849, 16032),
    "ES-AR": ("Aragón", 28.7, 35814, 30502, 2886, 2426),
    "ES-AS": ("Asturias", 26.9, 21244, 17625, 1876, 1743),
    "ES-IB": ("Illes Balears", 10.2, 14144, 11828, 1358, 958),
    "ES-CN": ("Canarias", 18.8, 47190, 40762, 4464, 1964),
    "ES-CB": ("Cantabria", 26.2, 16803, 11224, 4877, 702),
    "ES-CL": ("Castilla y León", 44.1, 94666, 77790, 10697, 6179),
    "ES-CM": ("Castilla-La Mancha", 13.7, 28206, 24084, 2511, 1611),
    "ES-CT": ("Cataluña", 27.9, 235242, 188279, 29463, 17500),
    "ES-VC": ("Comunitat Valenciana", 30.2, 170249, 134076, 24965, 11208),
    "ES-EX": ("Extremadura", 20.1, 20161, 16921, 2125, 1115),
    "ES-GA": ("Galicia", 27.6, 62635, 50527, 6775, 5333),
    "ES-MD": ("Comunidad de Madrid", 44.2, 341682, 261707, 61238, 18737),
    "ES-MC": ("Región de Murcia", 27.6, 52239, 42936, 6491, 2812),
    "ES-NC": ("Comunidad Foral de Navarra", 35.4, 24765, 18695, 4395, 1675),
    "ES-PV": ("País Vasco", 32.7, 63919, 53552, 5957, 4410),
    "ES-RI": ("La Rioja", 15.4, 4818, 3905, 589, 324),
}

UNIVERSITY_COUNTS_2024_25: dict[str, tuple[str, int, int, int]] = {
    "ES-AN": ("Andalucía", 12, 10, 2), "ES-AR": ("Aragón", 2, 1, 1),
    "ES-AS": ("Asturias", 1, 1, 0), "ES-IB": ("Illes Balears", 1, 1, 0),
    "ES-CN": ("Canarias", 6, 2, 4), "ES-CB": ("Cantabria", 2, 1, 1),
    "ES-CL": ("Castilla y León", 9, 4, 5), "ES-CM": ("Castilla-La Mancha", 1, 1, 0),
    "ES-CT": ("Cataluña", 12, 7, 5), "ES-VC": ("Comunitat Valenciana", 9, 5, 4),
    "ES-EX": ("Extremadura", 1, 1, 0), "ES-GA": ("Galicia", 4, 3, 1),
    "ES-MD": ("Comunidad de Madrid", 19, 6, 13), "ES-MC": ("Región de Murcia", 3, 2, 1),
    "ES-NC": ("Comunidad Foral de Navarra", 2, 1, 1), "ES-PV": ("País Vasco", 4, 1, 3),
    "ES-RI": ("La Rioja", 2, 1, 1),
}


def _university_series(id_: str, name: str, description: str, unit: str, decimals: int,
                       territory_points: dict[str, tuple[str, list[tuple[str, float | int]]]],
                       *, status: str = "Oficial (último dato provisional)") -> dict[str, Any]:
    series = series_base(
        id_=id_, name=name, description=description, unit=unit, frequency="Curso académico",
        source="SIIU · Ministerio de Ciencia, Innovación y Universidades",
        source_url=UNIVERSITY_REPORT_URL, status=status, decimals=decimals,
    )
    series["territories"] = order_territories([
        {"code": code, "name": label, "points": make_points(points)}
        for code, (label, points) in territory_points.items()
    ])
    return series


def collect_university_system(local_dir: Path | None = None) -> tuple[list[dict[str, Any]], SourceCheck]:
    """Indicadores del informe oficial Datos y cifras del SUE 2025-2026."""
    output: list[dict[str, Any]] = []
    national_latest = {
        "total": 1827272, "degree": 1416388, "master": 313189, "doctorate": 97695,
        "rate": 31.4,
    }
    national_2019 = {"total": 1626154, "degree": 1296379, "master": 237118, "doctorate": 92657}
    regional_indices = {"total": 2, "degree": 3, "master": 4, "doctorate": 5}
    labels = {
        "total": ("Estudiantes universitarios", "Estudiantes matriculados en grado, máster y doctorado."),
        "degree": ("Estudiantes de grado", "Estudiantes matriculados en enseñanzas universitarias de grado."),
        "master": ("Estudiantes de máster", "Estudiantes matriculados en másteres universitarios."),
        "doctorate": ("Estudiantes de doctorado", "Estudiantes matriculados en programas de doctorado."),
    }
    for key in ("total", "degree", "master", "doctorate"):
        territory_points: dict[str, tuple[str, list[tuple[str, int]]]] = {
            "ES": ("España", [("2019-20", national_2019[key]), ("2024-25", national_latest[key])])
        }
        idx = regional_indices[key]
        for code, row in UNIVERSITY_REGIONAL_2024_25.items():
            territory_points[code] = (row[0], [("2024-25", int(row[idx]))])
        name, desc = labels[key]
        output.append(_university_series(f"university_{key}_students", name, desc, "estudiantes", 0, territory_points))

    # Tasa neta regional y serie histórica nacional.
    rate_history = [
        ("2015-16", 31.5), ("2016-17", 32.0), ("2017-18", 32.1), ("2018-19", 31.8),
        ("2019-20", 31.5), ("2020-21", 32.1), ("2021-22", 31.9), ("2022-23", 31.4),
        ("2023-24", 31.1), ("2024-25", 31.4),
    ]
    rate_points: dict[str, tuple[str, list[tuple[str, float]]]] = {"ES": ("España", rate_history)}
    for code, row in UNIVERSITY_REGIONAL_2024_25.items():
        rate_points[code] = (row[0], [("2024-25", float(row[1]))])
    output.append(_university_series(
        "university_net_schooling_rate", "Tasa neta de escolarización universitaria",
        "Porcentaje de población de 18 a 24 años matriculada en grado o máster.", "%", 1, rate_points,
    ))

    grade_master_history = [
        ("2015-16", 1492741), ("2016-17", 1493395), ("2017-18", 1497545),
        ("2018-19", 1508295), ("2019-20", 1533497), ("2020-21", 1583260),
        ("2021-22", 1600490), ("2022-23", 1623748), ("2023-24", 1658389),
        ("2024-25", 1729577),
    ]
    output.append(_university_series(
        "university_grade_master_students", "Estudiantes universitarios de grado y máster",
        "Serie nacional conjunta de estudiantes matriculados en grado y máster.", "estudiantes", 0,
        {"ES": ("España", grade_master_history)},
    ))

    # Número de universidades por titularidad. El total nacional incluye dos universidades estatales.
    count_specs = [("total", "Universidades", 1), ("public", "Universidades públicas", 2), ("private", "Universidades privadas", 3)]
    national_counts = {"total": 92, "public": 50, "private": 42}
    for key, name, idx in count_specs:
        points: dict[str, tuple[str, list[tuple[str, int]]]] = {"ES": ("España", [("2024-25", national_counts[key])])}
        for code, row in UNIVERSITY_COUNTS_2024_25.items():
            points[code] = (row[0], [("2024-25", int(row[idx]))])
        output.append(_university_series(
            f"universities_{key}", name,
            f"Número de {name.lower()} en el Sistema Universitario Español.", "universidades", 0, points,
            status="Oficial",
        ))
    for key, label in (("public", "públicas"), ("private", "privadas")):
        points: dict[str, tuple[str, list[tuple[str, float]]]] = {
            "ES": ("España", [("2024-25", round(national_counts[key] / national_counts["total"] * 100, 2))])
        }
        idx = 2 if key == "public" else 3
        for code, row in UNIVERSITY_COUNTS_2024_25.items():
            points[code] = (row[0], [("2024-25", round(row[idx] / row[1] * 100, 2))])
        output.append(_university_series(
            f"universities_{key}_share", f"Porcentaje de universidades {label}",
            f"Porcentaje de universidades de titularidad {label[:-1]} sobre el total.", "%", 2, points,
            status="Calculado a partir de datos oficiales",
        ))

    # Titularidad del alumnado, disponible como serie nacional comparable.
    ownership = {
        "total": {
            "total": [("2019-20", 1626154), ("2024-25", 1827272)],
            "public": [("2019-20", 1306051), ("2024-25", 1330892)],
            "private": [("2019-20", 320103), ("2024-25", 496380)],
        },
        "degree": {
            "total": [("2019-20", 1296379), ("2024-25", 1416388)],
            "public": [("2019-20", 1079175), ("2024-25", 1092360)],
            "private": [("2019-20", 217204), ("2024-25", 324028)],
        },
        "master": {
            "total": [("2019-20", 237118), ("2024-25", 313189)],
            "public": [("2019-20", 139338), ("2024-25", 147501)],
            "private": [("2019-20", 97780), ("2024-25", 165688)],
        },
        "doctorate": {
            "total": [("2019-20", 92657), ("2024-25", 97695)],
            "public": [("2019-20", 87538), ("2024-25", 91031)],
            "private": [("2019-20", 5119), ("2024-25", 6664)],
        },
    }
    student_labels = {"total": "universitarios", "degree": "de grado", "master": "de máster", "doctorate": "de doctorado"}
    alumnado_labels = {"total": "universitario", "degree": "de grado", "master": "de máster", "doctorate": "de doctorado"}
    for level, data in ownership.items():
        for ownership_key in ("public", "private"):
            label = "públicas" if ownership_key == "public" else "privadas"
            output.append(_university_series(
                f"university_{level}_students_{ownership_key}",
                f"Estudiantes {student_labels[level]} en universidades {label}",
                f"Estudiantes {student_labels[level]} matriculados en universidades de titularidad {label[:-1]}.",
                "estudiantes", 0, {"ES": ("España", data[ownership_key])},
            ))
        total_map = dict(data["total"]); public_map = dict(data["public"]); private_map = dict(data["private"])
        pub_share = [(p, round(public_map[p] / total_map[p] * 100, 2)) for p in total_map]
        output.append(_university_series(
            f"university_{level}_public_share", f"Cuota pública del alumnado {alumnado_labels[level]}",
            f"Porcentaje del alumnado {alumnado_labels[level]} matriculado en universidades públicas.",
            "%", 2, {"ES": ("España", pub_share)},
        ))
        pri_share = [(p, round(private_map[p] / total_map[p] * 100, 2)) for p in total_map]
        output.append(_university_series(
            f"university_{level}_private_share", f"Cuota privada del alumnado {alumnado_labels[level]}",
            f"Porcentaje del alumnado {alumnado_labels[level]} matriculado en universidades privadas.",
            "%", 2, {"ES": ("España", pri_share)},
        ))

    # Controles internos de consistencia de las cifras publicadas.
    for level, data in ownership.items():
        for period, total in data["total"]:
            pub = dict(data["public"])[period]; pri = dict(data["private"])[period]
            if pub + pri != total:
                raise RuntimeError(f"Titularidad universitaria inconsistente: {level} {period}")
    if sum(row[1] for row in UNIVERSITY_COUNTS_2024_25.values()) + 2 != 92:
        raise RuntimeError("El total de universidades no coincide con el informe oficial")

    payload = repr((UNIVERSITY_REGIONAL_2024_25, UNIVERSITY_COUNTS_2024_25, ownership)).encode("utf-8")
    check = SourceCheck(
        "university_system", True, datetime.now(MADRID).isoformat(timespec="seconds"),
        f"{len(output)} series; curso 2024-25; cifras nacionales y autonómicas",
        sha256(payload), UNIVERSITY_REPORT_URL,
    )
    return output, check


def infer_category(series_id: str) -> str:
    if series_id.startswith(("education_", "university_", "universities_")): return "Enseñanza"
    if series_id.startswith(("surgical_", "consultation_", "hospital_", "primary_care_")): return "Sanidad"
    if series_id.startswith("dependency_"): return "Dependencia"
    if series_id.startswith("residential_"): return "Residencias"
    if series_id.startswith(("pension", "retirement_")): return "Pensiones y jubilación"
    if series_id.startswith("imv_"): return "Economía y empleo"
    if series_id.startswith("house_"): return "Vivienda"
    return "Otros indicadores"
